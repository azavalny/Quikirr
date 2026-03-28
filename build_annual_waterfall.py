"""
Build Annual_Waterfall_Output.xlsx from Source.xlsx (sheet 'Source Data').
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SOURCE_SHEET = "Source Data"
OUTPUT_SHEET = "Annual Waterfall ($ Actuals)"
ORIGINAL_SHEET = "Original Data"
DEFAULT_SOURCE = Path(__file__).resolve().parent / "Source.xlsx"
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "Annual_Waterfall_Output.xlsx"

BLUE = "1F4E79"
WHITE = "FFFFFF"
BEIGE = "F2E4D8"
GREY_HATCH_FG = "D9D9D9"
GREY_HATCH_BG = "F2F2F2"

THIN = Side(style="thin", color="000000")

ACCOUNTING_FMT = '_("$"* #,##0_);_("$"* (#,##0);_("$"* "-"??_);_(@_)'
PCT_FMT = "0%"
PCT_PARENS_NEG_FMT = '0%;(0%)'
CHURN_COUNT_FMT = '(#,##0)'
MRR_TO_ARR = 12


def _is_excel_error_value(v: Any) -> bool:
    if isinstance(v, str):
        t = v.strip()
        return bool(t.startswith("#") and t.endswith("!"))
    return False


def _is_customer_label_row(label: Any) -> bool:
    if label is None:
        return False
    if isinstance(label, str):
        s = label.strip()
        if not s or _is_excel_error_value(s):
            return False
        if s in "-–—":
            return False
        return True
    return isinstance(label, (int, float))


def _to_float(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if _is_excel_error_value(v):
        return 0.0
    if isinstance(v, str) and v.strip() in "-–—":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _sanitize_copy_value(v: Any) -> Any:
    if _is_excel_error_value(v):
        return None
    return v


def _display_currency(val: Any) -> Any:
    if val is None or isinstance(val, str):
        return val
    if isinstance(val, (int, float)):
        return int(round(float(val)))
    return val


def _display_pct(val: Any) -> Any:
    if val is None or isinstance(val, str):
        return val
    if isinstance(val, (int, float)):
        return round(float(val), 8)
    return val


def _parse_header_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def find_table_bounds(ws) -> tuple[int, int, int]:
    """Return (header_row, customer_col, first_date_col)."""
    best: tuple[int, int, int] | None = None
    for r in range(1, min(ws.max_row, 50) + 1):
        date_cols: list[int] = []
        cust_col: int | None = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and re.search(r"customer", v, re.I):
                cust_col = c
            if _parse_header_date(v) is not None:
                date_cols.append(c)
        if len(date_cols) >= 3 and cust_col is not None:
            if date_cols[0] > cust_col:
                return r, cust_col, date_cols[0]
        if len(date_cols) >= 3:
            first_d = min(date_cols)
            cc = cust_col if cust_col is not None and cust_col < first_d else first_d - 1
            if cc >= 1:
                return r, cc, first_d
    raise ValueError("Could not locate header row with dates and customer column.")


def year_end_column_index(header_row: int, ws) -> dict[int, int]:
    """Map calendar year -> 1-based column index (last month in year, prefer Dec)."""
    _, _, first_dc = find_table_bounds(ws)
    by_year: dict[int, list[tuple[int, date]]] = {}
    for c in range(first_dc, ws.max_column + 1):
        d = _parse_header_date(ws.cell(header_row, c).value)
        if d is None:
            continue
        by_year.setdefault(d.year, []).append((c, d))
    chosen: dict[int, int] = {}
    for y, lst in by_year.items():
        dec = [t for t in lst if t[1].month == 12 and t[1].day == 31]
        if dec:
            chosen[y] = max(dec, key=lambda t: t[1])[0]
        else:
            chosen[y] = max(lst, key=lambda t: t[1])[0]
    return chosen


@dataclass
class YearMetrics:
    bop_arr: float | None
    new_arr: float | None
    upsell: float | None
    downsell: float | None
    churn_arr: float | None
    eop_arr: float
    bop_cust: int | None
    new_cust: int | None
    churn_cust: int | None
    eop_cust: int
    needs_prior: bool


def compute_metrics_for_year(
    prev: list[float] | None, curr: list[float]
) -> YearMetrics:
    if prev is None:
        curr_pad = list(curr)
        eop_arr = float(sum(curr_pad))
        eop_cust = sum(1 for x in curr_pad if x > 0)
        return YearMetrics(
            None,
            None,
            None,
            None,
            None,
            eop_arr,
            None,
            None,
            None,
            eop_cust,
            False,
        )

    n = max(len(prev), len(curr))
    prev = list(prev) + [0.0] * (n - len(prev))
    curr = list(curr) + [0.0] * (n - len(curr))

    eop_arr = float(sum(curr))
    eop_cust = sum(1 for x in curr if x > 0)

    new_arr = 0.0
    upsell = 0.0
    downsell = 0.0
    churn_arr = 0.0
    new_cust = 0
    churn_cust = 0

    for p, c in zip(prev, curr):
        if p <= 0 and c > 0:
            new_arr += c
            new_cust += 1
        elif p > 0 and c == 0:
            churn_arr -= p
            churn_cust += 1
        elif p > 0 and c > 0:
            d = c - p
            if d > 0:
                upsell += d
            elif d < 0:
                downsell += d

    bop_arr = sum(prev)
    bop_cust = sum(1 for x in prev if x > 0)

    return YearMetrics(
        bop_arr,
        new_arr,
        upsell,
        downsell,
        churn_arr,
        eop_arr,
        bop_cust,
        new_cust,
        churn_cust,
        eop_cust,
        True,
    )


def read_customer_vectors(ws, header_row: int, cust_col: int, date_col: int) -> list[float]:
    vecs: list[float] = []
    for r in range(header_row + 1, ws.max_row + 1):
        label = ws.cell(r, cust_col).value
        if not _is_customer_label_row(label):
            continue
        vecs.append(_to_float(ws.cell(r, date_col).value) * MRR_TO_ARR)
    return vecs


def collect_snapshots(ws) -> tuple[int, int, dict[int, list[float]]]:
    header_row, cust_col, _ = find_table_bounds(ws)
    y2c = year_end_column_index(header_row, ws)
    years = sorted(y2c.keys())
    snaps: dict[int, list[float]] = {}
    for y in years:
        snaps[y] = read_customer_vectors(ws, header_row, cust_col, y2c[y])
    return header_row, cust_col, snaps


def dec_label(y: int) -> str:
    return f"Dec-{str(y)[-2:]}"


def build_summary_values(years: list[int], snaps: dict[int, list[float]]) -> tuple[list[str], list[dict[str, Any]]]:
    """Row specs: list of dict dec_yy -> value or None (hatched) or 'DASH' for en-dash."""
    col_keys = [dec_label(y) for y in years]
    metrics: list[YearMetrics] = []
    for i, y in enumerate(years):
        prev = snaps[years[i - 1]] if i > 0 else None
        curr = snaps[y]
        m = compute_metrics_for_year(prev, curr)
        metrics.append(m)

    rows: list[dict[str, Any]] = []

    def row() -> dict[str, Any]:
        return {k: None for k in col_keys}

    # BoP ARR
    r = row()
    for i, y in enumerate(years):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            r[col_keys[i]] = metrics[i].bop_arr
    rows.append({"label": "BoP ARR", "cells": r})

    for key, attr in [
        ("New", "new_arr"),
        ("Upsell", "upsell"),
        ("(Downsell)", "downsell"),
        ("(Churn)", "churn_arr"),
    ]:
        r = row()
        for i in range(len(years)):
            r[col_keys[i]] = "HATCH" if i == 0 else getattr(metrics[i], attr)
        rows.append({"label": key, "cells": r})

    r = row()
    for i in range(len(years)):
        r[col_keys[i]] = metrics[i].eop_arr
    rows.append({"label": "EoP ARR", "cells": r, "eop": True})

    # % Growth
    r = row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            b, e = metrics[i].bop_arr, metrics[i].eop_arr
            r[col_keys[i]] = (e / b - 1) if b and b != 0 else None
    rows.append({"label": "% Growth", "cells": r, "italic": True, "pct": True})

    # Net New ARR (EoP − BoP matches template; float bucket sum can differ by cents)
    r = row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            m = metrics[i]
            b = metrics[i - 1].eop_arr
            r[col_keys[i]] = m.eop_arr - b
    rows.append({"label": "Net New ARR", "cells": r, "italic": True, "currency": True})

    r = row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            m = metrics[i]
            r[col_keys[i]] = (m.new_arr or 0) + (m.upsell or 0)
    rows.append({"label": "New + Upsell", "cells": r, "italic": True, "currency": True})

    r = row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            m = metrics[i]
            r[col_keys[i]] = (m.downsell or 0) + (m.churn_arr or 0)
    rows.append({"label": "Downsell + Churn", "cells": r, "italic": True, "currency": True})

    for label, fn in [
        (
            "Gross Retention %",
            lambda m: ((m.bop_arr or 0) + (m.downsell or 0) + (m.churn_arr or 0)) / m.bop_arr
            if m.bop_arr
            else None,
        ),
        (
            "Loss-Only Retention %",
            lambda m: ((m.bop_arr or 0) + (m.churn_arr or 0)) / m.bop_arr if m.bop_arr else None,
        ),
        (
            "Net Retention %",
            lambda m: ((m.bop_arr or 0) + (m.upsell or 0) + (m.downsell or 0) + (m.churn_arr or 0))
            / m.bop_arr
            if m.bop_arr
            else None,
        ),
    ]:
        r = row()
        for i in range(len(years)):
            r[col_keys[i]] = "HATCH" if i == 0 else fn(metrics[i])
        rows.append({"label": label, "cells": r, "beige": True, "pct": True})

    rows.append({"label": "__CW_HEADER__", "cells": {}})

    r = row()
    for i in range(len(years)):
        r[col_keys[i]] = "HATCH" if i == 0 else metrics[i].bop_cust
    rows.append({"label": "BoP Customers", "cells": r, "int": True})

    r = row()
    for i in range(len(years)):
        r[col_keys[i]] = "HATCH" if i == 0 else metrics[i].new_cust
    rows.append({"label": "New", "cells": r, "int": True})

    r = row()
    for i in range(len(years)):
        r[col_keys[i]] = "HATCH" if i == 0 else metrics[i].churn_cust
    rows.append({"label": "(Churn)", "cells": r, "int": True, "churn_count": True})

    r = row()
    for i in range(len(years)):
        r[col_keys[i]] = metrics[i].eop_cust
    rows.append({"label": "EoP Customers", "cells": r, "eop": True, "int": True})

    r = row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            m = metrics[i]
            r[col_keys[i]] = (m.new_cust or 0) - (m.churn_cust or 0)
    rows.append({"label": "Net Logos Added", "cells": r, "italic": True, "int": True})

    r = row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "DASH"
        else:
            bc, ec = metrics[i].bop_cust, metrics[i].eop_cust
            r[col_keys[i]] = (ec / bc - 1) if bc else "DASH"
    rows.append({"label": "Logo % Growth", "cells": r, "italic": True, "pct_or_dash": True})

    r = row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            m = metrics[i]
            bc, ch = m.bop_cust or 0, m.churn_cust or 0
            r[col_keys[i]] = (bc - ch) / bc if bc else None
    rows.append({"label": "Gross Logo Retention %", "cells": r, "beige": True, "pct": True})

    def avg_logo(m: YearMetrics) -> float | None:
        if m.eop_cust:
            return m.eop_arr / m.eop_cust
        return None

    r = row()
    for i in range(len(years)):
        r[col_keys[i]] = "HATCH" if i == 0 else avg_logo(metrics[i])
    rows.append({"label": "Avg. Logo Size ($ Actuals)", "cells": r, "currency": True})

    r = row()
    for i in range(len(years)):
        if i < 2:
            r[col_keys[i]] = None
        else:
            a, b = avg_logo(metrics[i]), avg_logo(metrics[i - 1])
            r[col_keys[i]] = (a / b - 1) if a is not None and b not in (None, 0) else None
    rows.append({"label": "Avg Logo Size YoY Growth", "cells": r, "italic": True, "pct": True})

    def avg_new(m: YearMetrics) -> float | None:
        if m.new_cust and m.new_arr is not None:
            return m.new_arr / m.new_cust
        return None

    r = row()
    for i in range(len(years)):
        r[col_keys[i]] = "HATCH" if i == 0 else avg_new(metrics[i])
    rows.append({"label": "Avg. New Logo Size ($ Actuals)", "cells": r, "currency": True})

    r = row()
    for i in range(len(years)):
        if i < 2:
            r[col_keys[i]] = None
        else:
            a, b = avg_new(metrics[i]), avg_new(metrics[i - 1])
            r[col_keys[i]] = (a / b - 1) if a is not None and b not in (None, 0) else None
    rows.append(
        {
            "label": "Avg New Logo Size YoY Growth",
            "cells": r,
            "italic": True,
            "pct": True,
            "pct_parens_neg": True,
        }
    )

    def avg_churn(m: YearMetrics) -> float | None:
        if not m.needs_prior or m.churn_cust is None or m.churn_cust == 0:
            return None
        lost = -(m.churn_arr or 0)
        return lost / m.churn_cust

    r = row()
    for i in range(len(years)):
        r[col_keys[i]] = "HATCH" if i == 0 else avg_churn(metrics[i])
    rows.append({"label": "Avg. Churned Logo Size ($ Actuals)", "cells": r, "currency": True})

    r = row()
    for i in range(len(years)):
        if i < 2:
            r[col_keys[i]] = None
        else:
            a, b = avg_churn(metrics[i]), avg_churn(metrics[i - 1])
            r[col_keys[i]] = (a / b - 1) if a is not None and b not in (None, 0) else None
    rows.append(
        {
            "label": "Avg Churned Logo Size YoY Growth",
            "cells": r,
            "italic": True,
            "pct": True,
            "pct_parens_neg": True,
        }
    )

    return col_keys, rows


def copy_original_sheet(src_ws, dst_ws) -> None:
    for r in range(1, src_ws.max_row + 1):
        for c in range(1, src_ws.max_column + 1):
            dst_ws.cell(r, c, value=_sanitize_copy_value(src_ws.cell(r, c).value))


def write_summary_sheet(ws, col_keys: list[str], row_specs: list[dict[str, Any]]) -> None:
    fill_blue = PatternFill(fill_type="solid", fgColor=BLUE)
    fill_beige = PatternFill(fill_type="solid", fgColor=BEIGE)
    fill_hatch = PatternFill(
        patternType="lightDown",
        fgColor=GREY_HATCH_FG,
        bgColor=GREY_HATCH_BG,
    )
    font_white_bold = Font(bold=True, color=WHITE)
    font_bold = Font(bold=True)
    font_italic = Font(italic=True)

    ncols = len(col_keys)
    last_col_letter = get_column_letter(1 + ncols)

    ws.merge_cells(f"A1:{last_col_letter}1")
    c1 = ws["A1"]
    c1.value = "Annual Waterfall ($ Actuals)"
    c1.fill = fill_blue
    c1.font = font_white_bold
    c1.alignment = Alignment(horizontal="center")

    for j, ck in enumerate(col_keys, start=2):
        cell = ws.cell(2, j, value=ck)
        cell.fill = fill_blue
        cell.font = font_white_bold
        cell.alignment = Alignment(horizontal="right")
    ws.cell(2, 1).fill = fill_blue
    ws.cell(2, 1).value = None

    current_row = 3

    for spec in row_specs:
        if spec.get("label") == "__CW_HEADER__":
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=1 + ncols)
            h = ws.cell(current_row, 1, value="Customer Waterfall")
            h.fill = fill_blue
            h.font = font_white_bold
            for j in range(2, 2 + ncols):
                ws.cell(current_row, j).fill = fill_blue
            current_row += 1
            continue

        label = spec["label"]
        cells = spec["cells"]
        pct_fmt_row = PCT_PARENS_NEG_FMT if spec.get("pct_parens_neg") else PCT_FMT
        ws.cell(current_row, 1, value=label)
        ws.cell(current_row, 1).alignment = Alignment(horizontal="left")

        is_italic_label = spec.get("italic")
        if is_italic_label:
            ws.cell(current_row, 1).font = font_italic

        for j, ck in enumerate(col_keys, start=2):
            cell = ws.cell(current_row, j)
            val = cells.get(ck)
            if val == "HATCH":
                cell.value = None
                cell.fill = fill_hatch
            elif val == "DASH":
                cell.value = "–"
                cell.alignment = Alignment(horizontal="right")
                if is_italic_label:
                    cell.font = font_italic
            elif val is None:
                cell.value = None
            elif spec.get("pct_or_dash") and isinstance(val, (int, float)):
                cell.value = _display_pct(val)
                cell.number_format = pct_fmt_row
                cell.alignment = Alignment(horizontal="right")
                if is_italic_label:
                    cell.font = font_italic
            elif spec.get("pct"):
                cell.value = _display_pct(val)
                cell.number_format = pct_fmt_row
                cell.alignment = Alignment(horizontal="right")
                if is_italic_label:
                    cell.font = font_italic
            elif spec.get("currency"):
                cell.value = _display_currency(val)
                cell.number_format = ACCOUNTING_FMT
                cell.alignment = Alignment(horizontal="right")
                if is_italic_label:
                    cell.font = font_italic
            elif spec.get("int"):
                cell.value = val
                cell.alignment = Alignment(horizontal="right")
                if spec.get("churn_count") and isinstance(val, int):
                    cell.number_format = CHURN_COUNT_FMT
                else:
                    cell.number_format = "#,##0"
                if is_italic_label:
                    cell.font = font_italic
            else:
                cell.value = _display_currency(val)
                cell.number_format = ACCOUNTING_FMT
                cell.alignment = Alignment(horizontal="right")

            if spec.get("beige"):
                cell.fill = fill_beige
                ws.cell(current_row, 1).fill = fill_beige
                ws.cell(current_row, 1).font = font_bold
                cell.font = font_bold
                if isinstance(cell.value, (int, float)):
                    cell.number_format = pct_fmt_row

            if spec.get("eop"):
                cell.font = font_bold
                ws.cell(current_row, 1).font = font_bold
                cell.border = Border(top=THIN, bottom=THIN)
                ws.cell(current_row, 1).border = Border(top=THIN, bottom=THIN)
                if not spec.get("int"):
                    cell.number_format = ACCOUNTING_FMT

        current_row += 1

    ws.column_dimensions["A"].width = 36
    for j in range(2, 2 + ncols):
        ws.column_dimensions[get_column_letter(j)].width = 14


def assert_bridge(years: list[int], snaps: dict[int, list[float]], tol: float = 0.01) -> None:
    for i in range(1, len(years)):
        prev = snaps[years[i - 1]]
        curr = snaps[years[i]]
        m = compute_metrics_for_year(prev, curr)
        lhs = (
            (m.bop_arr or 0)
            + (m.new_arr or 0)
            + (m.upsell or 0)
            + (m.downsell or 0)
            + (m.churn_arr or 0)
        )
        rhs = m.eop_arr
        if abs(lhs - rhs) > tol:
            raise AssertionError(f"Bridge mismatch {years[i]}: {lhs} vs {rhs}")


def run(source: Path, output: Path, verify: bool = True) -> None:
    wb_in = load_workbook(source, data_only=True)
    if SOURCE_SHEET not in wb_in.sheetnames:
        raise KeyError(f"Missing sheet {SOURCE_SHEET!r}")
    ws_in = wb_in[SOURCE_SHEET]

    _, _, snaps = collect_snapshots(ws_in)
    years = sorted(snaps.keys())
    if len(years) < 1:
        raise ValueError("No year-end columns found.")

    if verify:
        assert_bridge(years, snaps)

    col_keys, row_specs = build_summary_values(years, snaps)

    wb_out = Workbook()
    ws_sum = wb_out.active
    ws_sum.title = OUTPUT_SHEET[:31]
    write_summary_sheet(ws_sum, col_keys, row_specs)

    ws_orig = wb_out.create_sheet(ORIGINAL_SHEET[:31])
    copy_original_sheet(ws_in, ws_orig)

    wb_out.save(output)
    wb_in.close()


def main() -> None:
    p = argparse.ArgumentParser(description="Build Annual Waterfall Excel from Source.xlsx")
    p.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    p.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    p.add_argument("--no-verify", action="store_true")
    args = p.parse_args()
    run(args.source, args.output, verify=not args.no_verify)
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
