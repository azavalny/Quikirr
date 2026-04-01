from __future__ import annotations

from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from ..config import INTERMEDIATE_SHEET, MRR_TO_ARR, ORIGINAL_SHEET
from ..source import SourceContext, dec_label, quarter_label
from ..styles import ACCOUNTING_FMT, FILL_BLUE, FONT_WHITE_BOLD


@dataclass
class IntermediateLayout:
    sheet_name: str
    cust_start_row: int
    cust_end_row: int
    rank_col: int
    year_cols: dict[int, int]
    annual_rows: dict[str, int]
    quarter_cols: dict
    quarterly_rows: dict[str, int]


SN_ORIG = f"'{ORIGINAL_SHEET}'"


class IntermediateCalculationsTab:
    title = INTERMEDIATE_SHEET

    def build(self, wb: Workbook, ctx: SourceContext) -> None:
        ws = wb.create_sheet(self.title[:31])
        years = ctx.years
        n = len(years)
        ds = ctx.data_start_row
        de = ctx.data_end_row
        num_rows = de - ds + 1

        year_cols: dict[int, int] = {y: i + 2 for i, y in enumerate(years)}
        rank_col = 2 + n

        # ---- Title ----
        ws["A1"].value = "Intermediate Calculations"
        ws["A1"].fill = FILL_BLUE
        ws["A1"].font = FONT_WHITE_BOLD
        ws["A1"].alignment = Alignment(horizontal="center")
        for c in range(2, rank_col + 1):
            ws.cell(1, c).fill = FILL_BLUE

        # ---- Section A: Per-Customer Year-End ARR ----
        hdr = ws.cell(2, 1, value="Customer")
        hdr.fill = FILL_BLUE
        hdr.font = FONT_WHITE_BOLD
        for y in years:
            cell = ws.cell(2, year_cols[y], value=dec_label(y))
            cell.fill = FILL_BLUE
            cell.font = FONT_WHITE_BOLD
            cell.alignment = Alignment(horizontal="right")
        rc = ws.cell(2, rank_col, value="Rank")
        rc.fill = FILL_BLUE
        rc.font = FONT_WHITE_BOLD
        rc.alignment = Alignment(horizontal="right")

        cust_start = 3
        cust_end = 2 + num_rows
        cust_cl = get_column_letter(ctx.cust_col)
        last_arr_cl = get_column_letter(year_cols[max(years)])

        for idx, src_r in enumerate(range(ds, de + 1)):
            r = cust_start + idx
            ws.cell(r, 1).value = f"={SN_ORIG}!${cust_cl}${src_r}"
            for y in years:
                src_ycl = ctx.year_col_letters[y]
                c = year_cols[y]
                ws.cell(r, c).value = (
                    f"={SN_ORIG}!${src_ycl}${src_r}*{MRR_TO_ARR}"
                )
                ws.cell(r, c).number_format = ACCOUNTING_FMT
            ws.cell(r, rank_col).value = (
                f"=IF(${last_arr_cl}{r}>10,"
                f"RANK(${last_arr_cl}{r},"
                f"${last_arr_cl}${cust_start}:${last_arr_cl}${cust_end},0)"
                f"+COUNTIF(${last_arr_cl}${cust_start}:"
                f"${last_arr_cl}{r},${last_arr_cl}{r})-1,"
                f'"N/A")'
            )

        # ---- Section B: Annual Waterfall Aggregates ----
        sb = cust_end + 2
        ws.cell(sb, 1, value="Annual Waterfall Aggregates").fill = FILL_BLUE
        ws.cell(sb, 1).font = FONT_WHITE_BOLD
        for y in years:
            cell = ws.cell(sb, year_cols[y], value=dec_label(y))
            cell.fill = FILL_BLUE
            cell.font = FONT_WHITE_BOLD
            cell.alignment = Alignment(horizontal="right")

        def _arr_rng(year: int) -> str:
            cl = get_column_letter(year_cols[year])
            return f"${cl}${cust_start}:${cl}${cust_end}"

        cur = sb + 1
        ar: dict[str, int] = {}

        ar["bop"] = cur
        ws.cell(cur, 1).value = "BoP ARR"
        for i, y in enumerate(years):
            c = year_cols[y]
            if i > 0:
                ws.cell(cur, c).value = f"=SUM({_arr_rng(years[i - 1])})"
            ws.cell(cur, c).number_format = ACCOUNTING_FMT
        cur += 1

        ar["new"] = cur
        ws.cell(cur, 1).value = "New"
        for i, y in enumerate(years):
            c = year_cols[y]
            if i > 0:
                pr, cr = _arr_rng(years[i - 1]), _arr_rng(y)
                ws.cell(cur, c).value = (
                    f"=SUMPRODUCT(({pr}<=0)*({cr}>0)*{cr})"
                )
            ws.cell(cur, c).number_format = ACCOUNTING_FMT
        cur += 1

        ar["up"] = cur
        ws.cell(cur, 1).value = "Upsell"
        for i, y in enumerate(years):
            c = year_cols[y]
            if i > 0:
                pr, cr = _arr_rng(years[i - 1]), _arr_rng(y)
                ws.cell(cur, c).value = (
                    f"=SUMPRODUCT(({pr}>0)*({cr}>0)"
                    f"*({cr}>{pr})*({cr}-{pr}))"
                )
            ws.cell(cur, c).number_format = ACCOUNTING_FMT
        cur += 1

        ar["ds"] = cur
        ws.cell(cur, 1).value = "(Downsell)"
        for i, y in enumerate(years):
            c = year_cols[y]
            if i > 0:
                pr, cr = _arr_rng(years[i - 1]), _arr_rng(y)
                ws.cell(cur, c).value = (
                    f"=SUMPRODUCT(({pr}>0)*({cr}>0)"
                    f"*({cr}<{pr})*({cr}-{pr}))"
                )
            ws.cell(cur, c).number_format = ACCOUNTING_FMT
        cur += 1

        ar["ch"] = cur
        ws.cell(cur, 1).value = "(Churn)"
        for i, y in enumerate(years):
            c = year_cols[y]
            if i > 0:
                pr, cr = _arr_rng(years[i - 1]), _arr_rng(y)
                ws.cell(cur, c).value = (
                    f"=-SUMPRODUCT(({pr}>0)*({cr}<=0)*{pr})"
                )
            ws.cell(cur, c).number_format = ACCOUNTING_FMT
        cur += 1

        ar["eop"] = cur
        ws.cell(cur, 1).value = "EoP ARR"
        for i, y in enumerate(years):
            c = year_cols[y]
            ws.cell(cur, c).value = f"=SUM({_arr_rng(y)})"
            ws.cell(cur, c).number_format = ACCOUNTING_FMT
        cur += 2

        ar["bopc"] = cur
        ws.cell(cur, 1).value = "BoP Customers"
        for i, y in enumerate(years):
            if i > 0:
                ws.cell(cur, year_cols[y]).value = (
                    f'=COUNTIF({_arr_rng(years[i - 1])},">0")'
                )
        cur += 1

        ar["newc"] = cur
        ws.cell(cur, 1).value = "New"
        for i, y in enumerate(years):
            if i > 0:
                pr, cr = _arr_rng(years[i - 1]), _arr_rng(y)
                ws.cell(cur, year_cols[y]).value = (
                    f"=SUMPRODUCT(({pr}<=0)*({cr}>0)*1)"
                )
        cur += 1

        ar["chc"] = cur
        ws.cell(cur, 1).value = "(Churn)"
        for i, y in enumerate(years):
            if i > 0:
                pr, cr = _arr_rng(years[i - 1]), _arr_rng(y)
                ws.cell(cur, year_cols[y]).value = (
                    f"=SUMPRODUCT(({pr}>0)*({cr}<=0)*1)"
                )
        cur += 1

        ar["eopc"] = cur
        ws.cell(cur, 1).value = "EoP Customers"
        for i, y in enumerate(years):
            ws.cell(cur, year_cols[y]).value = (
                f'=COUNTIF({_arr_rng(y)},">0")'
            )
        cur += 1

        # ---- Section C: Quarterly Waterfall Aggregates ----
        qks = ctx.quarter_keys
        nq = len(qks)
        qcols: dict[tuple[int, int], int] = {
            qk: qi + 2 for qi, qk in enumerate(qks)
        }

        sc = cur + 1
        ws.cell(sc, 1, value="Quarterly Waterfall Aggregates").fill = FILL_BLUE
        ws.cell(sc, 1).font = FONT_WHITE_BOLD
        for qk in qks:
            cell = ws.cell(sc, qcols[qk], value=quarter_label(qk[0], qk[1]))
            cell.fill = FILL_BLUE
            cell.font = FONT_WHITE_BOLD
            cell.alignment = Alignment(horizontal="right")

        def _qrng(key: tuple[int, int]) -> str:
            cl = ctx.quarter_col_letters[key]
            return f"{SN_ORIG}!${cl}${ds}:${cl}${de}"

        cur = sc + 1
        qr: dict[str, int] = {}

        qr["bop"] = cur
        ws.cell(cur, 1).value = "BoP ARR"
        for i, qk in enumerate(qks):
            c = qcols[qk]
            if i > 0:
                ws.cell(cur, c).value = (
                    f"=SUM({_qrng(qks[i - 1])})*{MRR_TO_ARR}"
                )
            ws.cell(cur, c).number_format = ACCOUNTING_FMT
        cur += 1

        qr["new"] = cur
        ws.cell(cur, 1).value = "New"
        for i, qk in enumerate(qks):
            c = qcols[qk]
            if i > 0:
                pr, cr = _qrng(qks[i - 1]), _qrng(qk)
                ws.cell(cur, c).value = (
                    f"=SUMPRODUCT(({pr}<=0)*({cr}>0)*{cr})*{MRR_TO_ARR}"
                )
            ws.cell(cur, c).number_format = ACCOUNTING_FMT
        cur += 1

        qr["up"] = cur
        ws.cell(cur, 1).value = "Upsell"
        for i, qk in enumerate(qks):
            c = qcols[qk]
            if i > 0:
                pr, cr = _qrng(qks[i - 1]), _qrng(qk)
                ws.cell(cur, c).value = (
                    f"=SUMPRODUCT(({pr}>0)*({cr}>0)"
                    f"*({cr}>{pr})*({cr}-{pr}))*{MRR_TO_ARR}"
                )
            ws.cell(cur, c).number_format = ACCOUNTING_FMT
        cur += 1

        qr["ds"] = cur
        ws.cell(cur, 1).value = "(Downsell)"
        for i, qk in enumerate(qks):
            c = qcols[qk]
            if i > 0:
                pr, cr = _qrng(qks[i - 1]), _qrng(qk)
                ws.cell(cur, c).value = (
                    f"=SUMPRODUCT(({pr}>0)*({cr}>0)"
                    f"*({cr}<{pr})*({cr}-{pr}))*{MRR_TO_ARR}"
                )
            ws.cell(cur, c).number_format = ACCOUNTING_FMT
        cur += 1

        qr["ch"] = cur
        ws.cell(cur, 1).value = "(Churn)"
        for i, qk in enumerate(qks):
            c = qcols[qk]
            if i > 0:
                pr, cr = _qrng(qks[i - 1]), _qrng(qk)
                ws.cell(cur, c).value = (
                    f"=-SUMPRODUCT(({pr}>0)*({cr}<=0)*{pr})*{MRR_TO_ARR}"
                )
            ws.cell(cur, c).number_format = ACCOUNTING_FMT
        cur += 1

        qr["eop"] = cur
        ws.cell(cur, 1).value = "EoP ARR"
        for i, qk in enumerate(qks):
            c = qcols[qk]
            ws.cell(cur, c).value = f"=SUM({_qrng(qk)})*{MRR_TO_ARR}"
            ws.cell(cur, c).number_format = ACCOUNTING_FMT
        cur += 2

        qr["bopc"] = cur
        ws.cell(cur, 1).value = "BoP Customers"
        for i, qk in enumerate(qks):
            if i > 0:
                ws.cell(cur, qcols[qk]).value = (
                    f'=COUNTIF({_qrng(qks[i - 1])},">0")'
                )
        cur += 1

        qr["newc"] = cur
        ws.cell(cur, 1).value = "New"
        for i, qk in enumerate(qks):
            if i > 0:
                pr, cr = _qrng(qks[i - 1]), _qrng(qk)
                ws.cell(cur, qcols[qk]).value = (
                    f"=SUMPRODUCT(({pr}<=0)*({cr}>0)*1)"
                )
        cur += 1

        qr["chc"] = cur
        ws.cell(cur, 1).value = "(Churn)"
        for i, qk in enumerate(qks):
            if i > 0:
                pr, cr = _qrng(qks[i - 1]), _qrng(qk)
                ws.cell(cur, qcols[qk]).value = (
                    f"=SUMPRODUCT(({pr}>0)*({cr}<=0)*1)"
                )
        cur += 1

        qr["eopc"] = cur
        ws.cell(cur, 1).value = "EoP Customers"
        for i, qk in enumerate(qks):
            ws.cell(cur, qcols[qk]).value = (
                f'=COUNTIF({_qrng(qk)},">0")'
            )

        # Column widths
        ws.column_dimensions["A"].width = 36
        max_col = max(rank_col, 1 + nq) if nq > 0 else rank_col
        for j in range(2, max_col + 1):
            ws.column_dimensions[get_column_letter(j)].width = 14

        ctx.intermediate = IntermediateLayout(
            sheet_name=self.title[:31],
            cust_start_row=cust_start,
            cust_end_row=cust_end,
            rank_col=rank_col,
            year_cols=year_cols,
            annual_rows=ar,
            quarter_cols=qcols,
            quarterly_rows=qr,
        )
