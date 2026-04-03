from __future__ import annotations

from collections.abc import Callable
from typing import Any

from openpyxl.utils import get_column_letter

from ..config import MRR_TO_ARR
from ..source import SourceContext, _to_float, dec_label, year_end_column_index
from .top_customers import _read_customers
from .top_sheet_common import (
    SN,
    TOP_N,
    add_conditional_formatting,
    apply_fmt,
    col_formats,
    col_layout,
    set_row_style,
    write_customer_formulas,
    write_header_block,
    write_other_row,
    write_sum_row,
    write_total_cohort_row,
)

SECTION_GAP = 2


def _arr_at(
    ctx: SourceContext, src_row: int, y: int, y2c: dict[int, int]
) -> float:
    c = y2c[y]
    return _to_float(ctx.ws_in.cell(src_row, c).value) * MRR_TO_ARR


def _pick_cohort(
    ctx: SourceContext,
    score_fn: Callable[[Any, float, float], float | None],
) -> list[Any]:
    years = ctx.years
    if len(years) < 2:
        return []
    y2c = year_end_column_index(ctx.header_row, ctx.ws_in)
    yp, yl = years[-2], years[-1]
    out: list[tuple[Any, float]] = []
    for cust in _read_customers(ctx):
        prev = _arr_at(ctx, cust.src_row, yp, y2c)
        last = _arr_at(ctx, cust.src_row, yl, y2c)
        s = score_fn(cust, prev, last)
        if s is not None:
            out.append((cust, s))
    out.sort(key=lambda t: (-t[1], t[0].src_row))
    return [c for c, _ in out[:TOP_N]]


def _score_new(_c: Any, prev: float, last: float) -> float | None:
    if prev <= 0 and last > 0 and last > 10:
        return last
    return None


def _score_upsell(_c: Any, prev: float, last: float) -> float | None:
    if prev > 0 and last > 0 and last > prev:
        return last - prev
    return None


def _score_downsell(_c: Any, prev: float, last: float) -> float | None:
    if prev > 10 and last > 0 and last < prev:
        return prev - last
    return None


def _score_churn(_c: Any, prev: float, last: float) -> float | None:
    if prev > 0 and last <= 0 and prev > 10:
        return prev
    return None


def _cohort_pct_header(years: list[int], category: str) -> str:
    return f"% of {category} {dec_label(years[-2])} - {dec_label(years[-1])}"


def _title_formula_main_table_date(n: int, prefix: str) -> str:
    last_hdr_col = get_column_letter(2 + n)
    return f'="{prefix} as of " & TEXT({last_hdr_col}3,"mmm-yy")'


def _append_one_cohort(
    ws,
    ctx: SourceContext,
    title_row: int,
    *,
    title_prefix: str,
    pct_category: str,
    total_label: str,
    other_prefix: str,
    rank_attr: str,
    score_attr: str,
    score_fn: Callable[[Any, float, float], float | None],
) -> int:
    years = ctx.years
    n = len(years)
    ic = ctx.intermediate
    rank_col_idx = getattr(ic, rank_attr, None)
    score_col_idx = getattr(ic, score_attr, None)

    if n < 2 or rank_col_idx is None or score_col_idx is None:
        return title_row

    layout = col_layout(n)
    total_cols = layout["since"]
    src_cols = {y: get_column_letter(c) for y, c in ic.year_cols.items()}
    cust_cl = "A"
    rank_cl = get_column_letter(rank_col_idx)
    score_cl = get_column_letter(score_col_idx)
    ds = ic.cust_start_row
    de = ic.cust_end_row

    data_start = title_row + 3
    top10 = _pick_cohort(ctx, score_fn)
    num_top = len(top10)

    other_label = (
        f'="Other {other_prefix} ("&COUNTIF({SN}!${score_cl}${ds}:'
        f'${score_cl}${de},">=0")-{num_top}&" Customers)"'
    )

    top10_end = data_start + min(TOP_N, len(top10)) - 1
    top10_total_row = top10_end + 1
    other_row = top10_total_row + 2
    total_row = other_row + 1
    top5_row = total_row + 2
    top10_row2 = top5_row + 1

    write_header_block(
        ws,
        title_row,
        years,
        layout,
        total_cols,
        title_a1=_title_formula_main_table_date(n, title_prefix),
        pct_arr_label=_cohort_pct_header(years, pct_category),
    )
    fmts = col_formats(layout, n)

    for rank_i, cust in enumerate(top10, start=1):
        row = data_start + rank_i - 1
        write_customer_formulas(
            ws,
            row,
            rank_i,
            cust.since,
            years,
            src_cols,
            layout,
            total_row,
            cust_cl,
            rank_cl,
            ds,
            de,
            block_data_start=data_start,
        )
        apply_fmt(ws, row, fmts)
        set_row_style(ws, row, total_cols)

    write_sum_row(
        ws,
        top10_total_row,
        "Total Top 10",
        data_start,
        top10_end,
        years,
        layout,
        total_row,
        bold=True,
        border=True,
        total_cols=total_cols,
    )
    apply_fmt(ws, top10_total_row, fmts)

    write_other_row(
        ws,
        other_row,
        other_label,
        top10_total_row,
        total_row,
        years,
        layout,
        total_cols,
    )
    apply_fmt(ws, other_row, fmts)

    write_total_cohort_row(
        ws,
        total_row,
        years,
        src_cols,
        layout,
        ds,
        de,
        total_cols,
        score_cl,
        total_label,
    )
    apply_fmt(ws, total_row, fmts)

    top5_end = data_start + min(5, len(top10)) - 1
    write_sum_row(
        ws,
        top5_row,
        "Total Top 5",
        data_start,
        top5_end,
        years,
        layout,
        total_row,
        total_cols=total_cols,
    )
    apply_fmt(ws, top5_row, fmts)

    write_sum_row(
        ws,
        top10_row2,
        "Total Top 10",
        data_start,
        top10_end,
        years,
        layout,
        total_row,
        total_cols=total_cols,
    )
    apply_fmt(ws, top10_row2, fmts)

    add_conditional_formatting(ws, layout, n, data_start, top10_row2)

    return top10_row2 + SECTION_GAP


def append_cohort_sections_to_top_customers_sheet(
    ws,
    ctx: SourceContext,
    first_title_row: int,
) -> None:
    r = first_title_row
    r = _append_one_cohort(
        ws,
        ctx,
        r,
        title_prefix="Top New Customers",
        pct_category="New",
        total_label="Total New Customers",
        other_prefix="New",
        rank_attr="rank_new_col",
        score_attr="score_new_col",
        score_fn=_score_new,
    )
    r = _append_one_cohort(
        ws,
        ctx,
        r,
        title_prefix="Top Upsell Customers",
        pct_category="Upsell",
        total_label="Total Upsell Customers",
        other_prefix="Upsell",
        rank_attr="rank_upsell_col",
        score_attr="score_upsell_col",
        score_fn=_score_upsell,
    )
    r = _append_one_cohort(
        ws,
        ctx,
        r,
        title_prefix="Top Downsell Customers",
        pct_category="Downsell",
        total_label="Total Downsell Customers",
        other_prefix="Downsell",
        rank_attr="rank_downsell_col",
        score_attr="score_downsell_col",
        score_fn=_score_downsell,
    )
    _append_one_cohort(
        ws,
        ctx,
        r,
        title_prefix="Top Churn Customers",
        pct_category="Churn",
        total_label="Total Churn Customers",
        other_prefix="Churn",
        rank_attr="rank_churn_col",
        score_attr="score_churn_col",
        score_fn=_score_churn,
    )
