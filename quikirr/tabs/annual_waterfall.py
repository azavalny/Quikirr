from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..config import MRR_TO_ARR, ORIGINAL_SHEET, OUTPUT_SHEET
from ..source import SourceContext, dec_label
from ..writer import write_summary_sheet


def build_formula_values(
    years: list[int],
    year_col_letters: dict[int, str],
    data_start_row: int,
    data_end_row: int,
    src_sheet: str = ORIGINAL_SHEET,
) -> tuple[list[str], list[dict[str, Any]]]:
    """Build row specs with Excel formula strings instead of hard-coded values."""
    col_keys = [dec_label(y) for y in years]
    sn = f"'{src_sheet}'"

    def _rng(year: int) -> str:
        cl = year_col_letters[year]
        return f"{sn}!${cl}${data_start_row}:${cl}${data_end_row}"

    def _oc(i: int) -> str:
        return get_column_letter(i + 2)

    rows: list[dict[str, Any]] = []

    def new_row() -> dict[str, Any]:
        return {k: None for k in col_keys}

    cur = 3

    R_BOP = cur; cur += 1
    r = new_row()
    for i, y in enumerate(years):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            r[col_keys[i]] = f"=SUM({_rng(years[i-1])})*{MRR_TO_ARR}"
    rows.append({"label": "BoP ARR", "cells": r})

    R_NEW = cur; cur += 1
    r = new_row()
    for i, y in enumerate(years):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            pr, cr = _rng(years[i - 1]), _rng(y)
            r[col_keys[i]] = f"=SUMPRODUCT(({pr}<=0)*({cr}>0)*{cr})*{MRR_TO_ARR}"
    rows.append({"label": "New", "cells": r})

    R_UP = cur; cur += 1
    r = new_row()
    for i, y in enumerate(years):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            pr, cr = _rng(years[i - 1]), _rng(y)
            r[col_keys[i]] = (
                f"=SUMPRODUCT(({pr}>0)*({cr}>0)*({cr}>{pr})"
                f"*({cr}-{pr}))*{MRR_TO_ARR}"
            )
    rows.append({"label": "Upsell", "cells": r})

    R_DS = cur; cur += 1
    r = new_row()
    for i, y in enumerate(years):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            pr, cr = _rng(years[i - 1]), _rng(y)
            r[col_keys[i]] = (
                f"=SUMPRODUCT(({pr}>0)*({cr}>0)*({cr}<{pr})"
                f"*({cr}-{pr}))*{MRR_TO_ARR}"
            )
    rows.append({"label": "(Downsell)", "cells": r})

    R_CH = cur; cur += 1
    r = new_row()
    for i, y in enumerate(years):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            pr, cr = _rng(years[i - 1]), _rng(y)
            r[col_keys[i]] = f"=-SUMPRODUCT(({pr}>0)*({cr}<=0)*{pr})*{MRR_TO_ARR}"
    rows.append({"label": "(Churn)", "cells": r})

    R_EOP = cur; cur += 1
    r = new_row()
    for i, y in enumerate(years):
        r[col_keys[i]] = f"=SUM({_rng(y)})*{MRR_TO_ARR}"
    rows.append({"label": "EoP ARR", "cells": r, "eop": True})

    cur += 1
    r = new_row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            oc = _oc(i)
            r[col_keys[i]] = f'=IF({oc}{R_BOP}=0,"",{oc}{R_EOP}/{oc}{R_BOP}-1)'
    rows.append({"label": "% Growth", "cells": r, "italic": True, "pct": True})

    cur += 1
    r = new_row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            oc = _oc(i)
            r[col_keys[i]] = f"={oc}{R_EOP}-{oc}{R_BOP}"
    rows.append({"label": "Net New ARR", "cells": r, "italic": True, "currency": True})

    cur += 1
    r = new_row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            oc = _oc(i)
            r[col_keys[i]] = f"={oc}{R_NEW}+{oc}{R_UP}"
    rows.append({"label": "New + Upsell", "cells": r, "italic": True, "currency": True})

    cur += 1
    r = new_row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            oc = _oc(i)
            r[col_keys[i]] = f"={oc}{R_DS}+{oc}{R_CH}"
    rows.append({"label": "Downsell + Churn", "cells": r, "italic": True, "currency": True})

    cur += 1
    r = new_row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            oc = _oc(i)
            r[col_keys[i]] = (
                f'=IF({oc}{R_BOP}=0,"",'
                f'({oc}{R_BOP}+{oc}{R_DS}+{oc}{R_CH})/{oc}{R_BOP})'
            )
    rows.append({"label": "Gross Retention %", "cells": r, "beige": True, "pct": True})

    cur += 1
    r = new_row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            oc = _oc(i)
            r[col_keys[i]] = (
                f'=IF({oc}{R_BOP}=0,"",'
                f'({oc}{R_BOP}+{oc}{R_CH})/{oc}{R_BOP})'
            )
    rows.append({"label": "Loss-Only Retention %", "cells": r, "beige": True, "pct": True})

    cur += 1
    r = new_row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            oc = _oc(i)
            r[col_keys[i]] = (
                f'=IF({oc}{R_BOP}=0,"",'
                f'({oc}{R_BOP}+{oc}{R_UP}+{oc}{R_DS}+{oc}{R_CH})/{oc}{R_BOP})'
            )
    rows.append({"label": "Net Retention %", "cells": r, "beige": True, "pct": True})

    rows.append({"label": "__CW_HEADER__", "cells": {}})
    cur += 1

    R_BOPC = cur; cur += 1
    r = new_row()
    for i, y in enumerate(years):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            r[col_keys[i]] = f'=COUNTIF({_rng(years[i-1])},">0")'
    rows.append({"label": "BoP Customers", "cells": r, "int": True})

    R_NEWC = cur; cur += 1
    r = new_row()
    for i, y in enumerate(years):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            pr, cr = _rng(years[i - 1]), _rng(y)
            r[col_keys[i]] = f"=SUMPRODUCT(({pr}<=0)*({cr}>0)*1)"
    rows.append({"label": "New", "cells": r, "int": True})

    R_CHC = cur; cur += 1
    r = new_row()
    for i, y in enumerate(years):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            pr, cr = _rng(years[i - 1]), _rng(y)
            r[col_keys[i]] = f"=SUMPRODUCT(({pr}>0)*({cr}<=0)*1)"
    rows.append({"label": "(Churn)", "cells": r, "int": True, "churn_count": True})

    R_EOPC = cur; cur += 1
    r = new_row()
    for i, y in enumerate(years):
        r[col_keys[i]] = f'=COUNTIF({_rng(y)},">0")'
    rows.append({"label": "EoP Customers", "cells": r, "eop": True, "int": True})

    cur += 1
    r = new_row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            oc = _oc(i)
            r[col_keys[i]] = f"={oc}{R_NEWC}-{oc}{R_CHC}"
    rows.append({"label": "Net Logos Added", "cells": r, "italic": True, "int": True})

    cur += 1
    r = new_row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "DASH"
        else:
            oc = _oc(i)
            r[col_keys[i]] = f'=IF({oc}{R_BOPC}=0,"–",{oc}{R_EOPC}/{oc}{R_BOPC}-1)'
    rows.append({"label": "Logo % Growth", "cells": r, "italic": True, "pct_or_dash": True, "pct": True})

    cur += 1
    r = new_row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            oc = _oc(i)
            r[col_keys[i]] = (
                f'=IF({oc}{R_BOPC}=0,"",'
                f'({oc}{R_BOPC}-{oc}{R_CHC})/{oc}{R_BOPC})'
            )
    rows.append({"label": "Gross Logo Retention %", "cells": r, "beige": True, "pct": True})

    R_AVGL = cur; cur += 1
    r = new_row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            oc = _oc(i)
            r[col_keys[i]] = f'=IF({oc}{R_EOPC}=0,"",{oc}{R_EOP}/{oc}{R_EOPC})'
    rows.append({"label": "Avg. Logo Size ($ Actuals)", "cells": r, "currency": True})

    cur += 1
    r = new_row()
    for i in range(len(years)):
        if i < 2:
            r[col_keys[i]] = None
        else:
            oc, poc = _oc(i), _oc(i - 1)
            r[col_keys[i]] = (
                f'=IF(OR({poc}{R_AVGL}=0,{poc}{R_AVGL}=""),"",'
                f'{oc}{R_AVGL}/{poc}{R_AVGL}-1)'
            )
    rows.append({"label": "Avg Logo Size YoY Growth", "cells": r, "italic": True, "pct": True})

    R_AVGNL = cur; cur += 1
    r = new_row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            oc = _oc(i)
            r[col_keys[i]] = f'=IF({oc}{R_NEWC}=0,"",{oc}{R_NEW}/{oc}{R_NEWC})'
    rows.append({"label": "Avg. New Logo Size ($ Actuals)", "cells": r, "currency": True})

    cur += 1
    r = new_row()
    for i in range(len(years)):
        if i < 2:
            r[col_keys[i]] = None
        else:
            oc, poc = _oc(i), _oc(i - 1)
            r[col_keys[i]] = (
                f'=IF(OR({poc}{R_AVGNL}=0,{poc}{R_AVGNL}=""),"",'
                f'{oc}{R_AVGNL}/{poc}{R_AVGNL}-1)'
            )
    rows.append(
        {
            "label": "Avg New Logo Size YoY Growth",
            "cells": r,
            "italic": True,
            "pct": True,
            "pct_parens_neg": True,
        }
    )

    R_AVGCL = cur; cur += 1
    r = new_row()
    for i in range(len(years)):
        if i == 0:
            r[col_keys[i]] = "HATCH"
        else:
            oc = _oc(i)
            r[col_keys[i]] = f'=IF({oc}{R_CHC}=0,"",-{oc}{R_CH}/{oc}{R_CHC})'
    rows.append({"label": "Avg. Churned Logo Size ($ Actuals)", "cells": r, "currency": True})

    cur += 1
    r = new_row()
    for i in range(len(years)):
        if i < 2:
            r[col_keys[i]] = None
        else:
            oc, poc = _oc(i), _oc(i - 1)
            r[col_keys[i]] = (
                f'=IF(OR({poc}{R_AVGCL}=0,{poc}{R_AVGCL}=""),"",'
                f'{oc}{R_AVGCL}/{poc}{R_AVGCL}-1)'
            )
    rows.append(
        {
            "label": "Avg Churned Logo Size YoY Growth",
            "cells": r,
            "italic": True,
            "pct": True,
            "pct_parens_neg": True,
        }
    )

    return col_keys, rows


class AnnualWaterfallTab:
    title = OUTPUT_SHEET

    def build(self, wb: Workbook, ctx: SourceContext) -> None:
        col_keys, row_specs = build_formula_values(
            ctx.years,
            ctx.year_col_letters,
            ctx.data_start_row,
            ctx.data_end_row,
        )
        ws = wb.active if wb.active.title == "Sheet" else wb.create_sheet()
        ws.title = self.title[:31]
        write_summary_sheet(ws, self.title, col_keys, row_specs)
