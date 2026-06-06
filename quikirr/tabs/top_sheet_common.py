from __future__ import annotations

from typing import Any

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border
from openpyxl.utils import get_column_letter

from ..config import INTERMEDIATE_SHEET
from ..source import dec_label
from ..styles import (
    ACCOUNTING_FMT,
    FILL_BLUE,
    FILL_GREEN_COND,
    FILL_RED_COND,
    FONT_BOLD,
    FONT_WHITE_BOLD,
    MULT_FMT,
    PCT_FMT,
    PCT_PARENS_NEG_FMT,
    THIN,
)

TOP_N = 10
DATA_START = 4
SN = f"'{INTERMEDIATE_SHEET}'"


def col_layout(n: int) -> dict[str, Any]:
    arr_start = 3
    cagr = arr_start + n
    pct_arr = cagr + 1
    pct_cum = pct_arr + 1
    dg_start = pct_cum + 1
    pct_of_g = dg_start + (n - 1)
    mult = pct_of_g + 1
    pg_start = mult + 1
    since = pg_start + (n - 1)
    total = since
    return dict(
        arr_start=arr_start,
        cagr=cagr,
        pct_arr=pct_arr,
        pct_cum=pct_cum,
        dg_start=dg_start,
        pct_of_g=pct_of_g,
        mult=mult,
        pg_start=pg_start,
        since=since,
        total=total,
    )


def merge_super(ws, row: int, start: int, end: int, value: str) -> None:
    first = ws.cell(row, start)
    first.value = value
    first.alignment = Alignment(horizontal="center")
    for c in range(start + 1, end + 1):
        ws.cell(row, c).value = None


def write_header_block(
    ws,
    top_row: int,
    years: list[int],
    layout: dict,
    total_cols: int,
    *,
    title_a1: str,
    pct_arr_label: str | None = None,
) -> None:
    n = len(years)
    first_lbl = dec_label(years[0])
    last_lbl = dec_label(years[-1])
    r1, r2, r3 = top_row, top_row + 1, top_row + 2
    pct_hdr = pct_arr_label if pct_arr_label is not None else f"% of {last_lbl} ARR"

    title = ws.cell(r1, 1)
    title.value = title_a1
    title.fill = FILL_BLUE
    title.font = FONT_WHITE_BOLD
    title.alignment = Alignment(horizontal="center")
    for c in range(2, total_cols + 1):
        ws.cell(r1, c).fill = FILL_BLUE

    for c in range(1, total_cols + 1):
        for rr in (r2, r3):
            cell = ws.cell(rr, c)
            cell.fill = FILL_BLUE
            cell.font = FONT_WHITE_BOLD

    ws.cell(r2, 1).value = "($ Actuals)"
    ws.cell(r2, 1).alignment = Alignment(horizontal="left")

    merge_super(ws, r2, layout["arr_start"], layout["pct_cum"], "ARR")
    merge_super(ws, r2, layout["dg_start"], layout["mult"], "$ Growth")
    merge_super(ws, r2, layout["pg_start"], layout["since"] - 1, "% Growth")
    ws.cell(r2, layout["since"]).value = ""

    col = 1
    ws.cell(r3, col).value = "Rank"
    col += 1
    ws.cell(r3, col).value = "Customer"
    col += 1

    for y in years:
        ws.cell(r3, col).value = dec_label(y)
        ws.cell(r3, col).alignment = Alignment(horizontal="right")
        col += 1

    ws.cell(r3, col).value = f"{first_lbl} - {last_lbl} CAGR"
    ws.cell(r3, col).alignment = Alignment(horizontal="right")
    col += 1
    ws.cell(r3, col).value = pct_hdr
    ws.cell(r3, col).alignment = Alignment(horizontal="right")
    col += 1
    ws.cell(r3, col).value = "% Total Cum."
    ws.cell(r3, col).alignment = Alignment(horizontal="right")
    col += 1

    for i in range(n - 1):
        ws.cell(r3, col).value = f"{dec_label(years[i])} - {dec_label(years[i + 1])}"
        ws.cell(r3, col).alignment = Alignment(horizontal="right")
        col += 1
    ws.cell(r3, col).value = f"% of {dec_label(years[-2])} - {last_lbl} Growth"
    ws.cell(r3, col).alignment = Alignment(horizontal="right")
    col += 1
    ws.cell(r3, col).value = "Mult. of Initial"
    ws.cell(r3, col).alignment = Alignment(horizontal="right")
    col += 1

    for i in range(n - 1):
        ws.cell(r3, col).value = f"{dec_label(years[i])} - {dec_label(years[i + 1])}"
        ws.cell(r3, col).alignment = Alignment(horizontal="right")
        col += 1

    ws.cell(r3, col).value = "Customer Since"
    ws.cell(r3, col).alignment = Alignment(horizontal="right")


def write_headers(
    ws,
    years: list[int],
    layout: dict,
    total_cols: int,
    *,
    title_a1: str,
    pct_arr_label: str | None = None,
) -> None:
    write_header_block(
        ws, 1, years, layout, total_cols, title_a1=title_a1, pct_arr_label=pct_arr_label
    )


def col_formats(layout: dict, n: int) -> list[tuple[int, str]]:
    fmts: list[tuple[int, str]] = []
    for i in range(n):
        fmts.append((layout["arr_start"] + i, ACCOUNTING_FMT))
    fmts.append((layout["cagr"], PCT_FMT))
    fmts.append((layout["pct_arr"], PCT_FMT))
    fmts.append((layout["pct_cum"], PCT_FMT))
    for i in range(n - 1):
        fmts.append((layout["dg_start"] + i, ACCOUNTING_FMT))
    fmts.append((layout["pct_of_g"], PCT_FMT))
    fmts.append((layout["mult"], MULT_FMT))
    for i in range(n - 1):
        fmts.append((layout["pg_start"] + i, PCT_PARENS_NEG_FMT))
    return fmts


def apply_fmt(ws, row: int, fmts: list[tuple[int, str]]) -> None:
    for col, fmt in fmts:
        ws.cell(row, col).number_format = fmt


def set_row_style(
    ws, row: int, total_cols: int, bold: bool = False, border: bool = False
) -> None:
    for c in range(1, total_cols + 1):
        cell = ws.cell(row, c)
        cell.alignment = Alignment(horizontal="right" if c > 2 else "left")
        if bold:
            cell.font = FONT_BOLD
        if border:
            cell.border = Border(top=THIN, bottom=THIN)


def write_customer_formulas(
    ws,
    row: int,
    rank: int,
    customer_since: str,
    years: list[int],
    src_cols: dict[int, str],
    layout: dict,
    total_row: int,
    cust_cl: str,
    rank_cl: str,
    ds: int,
    de: int,
    block_data_start: int = DATA_START,
) -> None:
    n = len(years)
    span = n - 1
    ws.cell(row, 1).value = rank
    ws.cell(row, 2).value = (
        f"=IFERROR(INDEX({SN}!${cust_cl}${ds}:${cust_cl}${de},"
        f"MATCH($A{row},{SN}!${rank_cl}${ds}:${rank_cl}${de},0),1),\"\")"
    )

    acl: list[str] = []
    for i, y in enumerate(years):
        col = layout["arr_start"] + i
        cl = get_column_letter(col)
        acl.append(cl)
        ws.cell(row, col).value = (
            f"=IFERROR(INDEX({SN}!${src_cols[y]}${ds}:${src_cols[y]}${de},"
            f"MATCH($A{row},{SN}!${rank_cl}${ds}:${rank_cl}${de},0),1),0)"
        )

    first, last = acl[0], acl[-1]
    ws.cell(row, layout["cagr"]).value = (
        f"=IF({first}{row}=0,0,({last}{row}/{first}{row})^(1/{span})-1)"
    )

    pct_cl = get_column_letter(layout["pct_arr"])
    ws.cell(row, layout["pct_arr"]).value = (
        f"=IF({last}{total_row}=0,0,{last}{row}/{last}{total_row})"
    )
    ws.cell(row, layout["pct_cum"]).value = (
        f"=SUM({pct_cl}${block_data_start}:{pct_cl}{row})"
    )

    dgcl: list[str] = []
    for i in range(n - 1):
        col = layout["dg_start"] + i
        cl = get_column_letter(col)
        dgcl.append(cl)
        ws.cell(row, col).value = f"={acl[i + 1]}{row}-{acl[i]}{row}"

    last_dg = dgcl[-1]
    ws.cell(row, layout["pct_of_g"]).value = (
        f"=IF({last_dg}{total_row}=0,0,{last_dg}{row}/{last_dg}{total_row})"
    )
    ws.cell(row, layout["mult"]).value = (
        f"=IF({first}{row}=0,0,{last}{row}/{first}{row})"
    )

    for i in range(n - 1):
        col = layout["pg_start"] + i
        ws.cell(row, col).value = (
            f"=IF({acl[i]}{row}=0,0,{acl[i + 1]}{row}/{acl[i]}{row}-1)"
        )

    ws.cell(row, layout["since"]).value = customer_since
    ws.cell(row, layout["since"]).alignment = Alignment(horizontal="right")


def section_row_layout(data_start: int) -> dict[str, int]:
    data_end = data_start + TOP_N - 1
    top10_total_row = data_end + 1
    other_row = top10_total_row + 2
    total_row = other_row + 1
    top5_row = total_row + 2
    top10_row2 = top5_row + 1
    return dict(
        data_end=data_end,
        top5_end=data_start + 5 - 1,
        top10_total_row=top10_total_row,
        other_row=other_row,
        total_row=total_row,
        top5_row=top5_row,
        top10_row2=top10_row2,
    )


def write_top_n_customer_rows(
    ws,
    data_start: int,
    customers: list,
    years: list[int],
    src_cols: dict[int, str],
    layout: dict,
    total_row: int,
    cust_cl: str,
    rank_cl: str,
    ds: int,
    de: int,
    fmts: list[tuple[int, str]],
    total_cols: int,
) -> None:
    for rank_i in range(1, TOP_N + 1):
        row = data_start + rank_i - 1
        since = customers[rank_i - 1].since if rank_i - 1 < len(customers) else ""
        write_customer_formulas(
            ws,
            row,
            rank_i,
            since,
            years,
            src_cols,
            layout,
            total_row,
            cust_cl,
            rank_cl,
            ds,
            de,
            block_data_start=data_start,
        )
        apply_fmt(ws, row, fmts)
        set_row_style(ws, row, total_cols)


def write_sum_row(
    ws,
    row: int,
    label: str,
    sum_start: int,
    sum_end: int,
    years: list[int],
    layout: dict,
    total_row: int,
    bold: bool = False,
    border: bool = False,
    total_cols: int = 0,
) -> None:
    n = len(years)
    span = n - 1
    ws.cell(row, 2).value = label

    acl: list[str] = []
    for i in range(n):
        col = layout["arr_start"] + i
        cl = get_column_letter(col)
        acl.append(cl)
        ws.cell(row, col).value = f"=SUM({cl}{sum_start}:{cl}{sum_end})"

    first, last = acl[0], acl[-1]
    ws.cell(row, layout["cagr"]).value = (
        f"=IF({first}{row}=0,0,({last}{row}/{first}{row})^(1/{span})-1)"
    )
    pct_cl = get_column_letter(layout["pct_arr"])
    ws.cell(row, layout["pct_arr"]).value = (
        f"=IF({last}{total_row}=0,0,{last}{row}/{last}{total_row})"
    )
    ws.cell(row, layout["pct_cum"]).value = (
        f"=SUM({pct_cl}{sum_start}:{pct_cl}{sum_end})"
    )

    dgcl: list[str] = []
    for i in range(n - 1):
        col = layout["dg_start"] + i
        cl = get_column_letter(col)
        dgcl.append(cl)
        ws.cell(row, col).value = f"=SUM({cl}{sum_start}:{cl}{sum_end})"

    last_dg = dgcl[-1]
    ws.cell(row, layout["pct_of_g"]).value = (
        f"=IF({last_dg}{total_row}=0,0,{last_dg}{row}/{last_dg}{total_row})"
    )
    ws.cell(row, layout["mult"]).value = (
        f"=IF({first}{row}=0,0,{last}{row}/{first}{row})"
    )

    for i in range(n - 1):
        col = layout["pg_start"] + i
        ws.cell(row, col).value = (
            f"=IF({acl[i]}{row}=0,0,{acl[i + 1]}{row}/{acl[i]}{row}-1)"
        )

    set_row_style(ws, row, total_cols, bold=bold, border=border)


def write_total_customers_row(
    ws,
    row: int,
    years: list[int],
    src_cols: dict[int, str],
    layout: dict,
    data_start_row: int,
    data_end_row: int,
    total_cols: int,
) -> None:
    n = len(years)
    span = n - 1
    ws.cell(row, 2).value = "Total Customers"

    acl: list[str] = []
    for i, y in enumerate(years):
        col = layout["arr_start"] + i
        cl = get_column_letter(col)
        acl.append(cl)
        rng = f"{SN}!${src_cols[y]}${data_start_row}:${src_cols[y]}${data_end_row}"
        ws.cell(row, col).value = f"=SUM({rng})"

    first, last = acl[0], acl[-1]
    ws.cell(row, layout["cagr"]).value = (
        f"=IF({first}{row}=0,0,({last}{row}/{first}{row})^(1/{span})-1)"
    )
    ws.cell(row, layout["pct_arr"]).value = (
        f"=IF({last}{row}=0,0,{last}{row}/{last}{row})"
    )
    ws.cell(row, layout["pct_cum"]).value = (
        f"=IF({last}{row}=0,0,{last}{row}/{last}{row})"
    )

    dgcl: list[str] = []
    for i in range(n - 1):
        col = layout["dg_start"] + i
        cl = get_column_letter(col)
        dgcl.append(cl)
        ws.cell(row, col).value = f"={acl[i + 1]}{row}-{acl[i]}{row}"

    last_dg = dgcl[-1]
    ws.cell(row, layout["pct_of_g"]).value = (
        f"=IF({last_dg}{row}=0,0,{last_dg}{row}/{last_dg}{row})"
    )
    ws.cell(row, layout["mult"]).value = (
        f"=IF({first}{row}=0,0,{last}{row}/{first}{row})"
    )

    for i in range(n - 1):
        col = layout["pg_start"] + i
        ws.cell(row, col).value = (
            f"=IF({acl[i]}{row}=0,0,{acl[i + 1]}{row}/{acl[i]}{row}-1)"
        )

    set_row_style(ws, row, total_cols, bold=True, border=True)


def write_total_cohort_row(
    ws,
    row: int,
    years: list[int],
    src_cols: dict[int, str],
    layout: dict,
    data_start_row: int,
    data_end_row: int,
    total_cols: int,
    score_col_letter: str,
    total_label: str,
) -> None:
    n = len(years)
    span = n - 1
    ws.cell(row, 2).value = total_label
    sc = score_col_letter
    ds, de = data_start_row, data_end_row

    acl: list[str] = []
    for i, y in enumerate(years):
        col = layout["arr_start"] + i
        cl = get_column_letter(col)
        acl.append(cl)
        ycl = src_cols[y]
        ws.cell(row, col).value = (
            f"=SUMPRODUCT(({SN}!${sc}${ds}:${sc}${de}>=0)*"
            f"({SN}!${ycl}${ds}:${ycl}${de}))"
        )

    first, last = acl[0], acl[-1]
    ws.cell(row, layout["cagr"]).value = (
        f"=IF({first}{row}=0,0,({last}{row}/{first}{row})^(1/{span})-1)"
    )
    ws.cell(row, layout["pct_arr"]).value = (
        f"=IF({last}{row}=0,0,{last}{row}/{last}{row})"
    )
    ws.cell(row, layout["pct_cum"]).value = (
        f"=IF({last}{row}=0,0,{last}{row}/{last}{row})"
    )

    dgcl: list[str] = []
    for i in range(n - 1):
        col = layout["dg_start"] + i
        cl = get_column_letter(col)
        dgcl.append(cl)
        ws.cell(row, col).value = f"={acl[i + 1]}{row}-{acl[i]}{row}"

    last_dg = dgcl[-1]
    ws.cell(row, layout["pct_of_g"]).value = (
        f"=IF({last_dg}{row}=0,0,{last_dg}{row}/{last_dg}{row})"
    )
    ws.cell(row, layout["mult"]).value = (
        f"=IF({first}{row}=0,0,{last}{row}/{first}{row})"
    )

    for i in range(n - 1):
        col = layout["pg_start"] + i
        ws.cell(row, col).value = (
            f"=IF({acl[i]}{row}=0,0,{acl[i + 1]}{row}/{acl[i]}{row}-1)"
        )

    set_row_style(ws, row, total_cols, bold=True, border=True)


def write_other_row(
    ws,
    row: int,
    label: str,
    top10_row: int,
    total_row: int,
    years: list[int],
    layout: dict,
    total_cols: int,
) -> None:
    n = len(years)
    span = n - 1
    ws.cell(row, 2).value = label

    acl: list[str] = []
    for i in range(n):
        col = layout["arr_start"] + i
        cl = get_column_letter(col)
        acl.append(cl)
        ws.cell(row, col).value = f"={cl}{total_row}-{cl}{top10_row}"

    first, last = acl[0], acl[-1]
    ws.cell(row, layout["cagr"]).value = (
        f"=IF({first}{row}=0,0,({last}{row}/{first}{row})^(1/{span})-1)"
    )
    ws.cell(row, layout["pct_arr"]).value = (
        f"=IF({last}{total_row}=0,0,{last}{row}/{last}{total_row})"
    )
    ws.cell(row, layout["pct_cum"]).value = ""

    dgcl: list[str] = []
    for i in range(n - 1):
        col = layout["dg_start"] + i
        cl = get_column_letter(col)
        dgcl.append(cl)
        ws.cell(row, col).value = f"={cl}{total_row}-{cl}{top10_row}"

    last_dg = dgcl[-1]
    ws.cell(row, layout["pct_of_g"]).value = (
        f"=IF({last_dg}{total_row}=0,0,{last_dg}{row}/{last_dg}{total_row})"
    )
    ws.cell(row, layout["mult"]).value = (
        f"=IF({first}{row}=0,0,{last}{row}/{first}{row})"
    )

    for i in range(n - 1):
        col = layout["pg_start"] + i
        ws.cell(row, col).value = (
            f"=IF({acl[i]}{row}=0,0,{acl[i + 1]}{row}/{acl[i]}{row}-1)"
        )

    set_row_style(ws, row, total_cols)


def add_conditional_formatting(
    ws, layout: dict, n: int, first_row: int, last_row: int
) -> None:
    for i in range(n - 1):
        cl = get_column_letter(layout["dg_start"] + i)
        rng = f"{cl}{first_row}:{cl}{last_row}"
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="greaterThan", formula=["0"], fill=FILL_GREEN_COND),
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="lessThan", formula=["0"], fill=FILL_RED_COND)
        )


def set_column_widths(ws, total_cols: int) -> None:
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    for j in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 14
