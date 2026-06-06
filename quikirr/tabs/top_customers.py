from __future__ import annotations

from dataclasses import dataclass
from datetime import date

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..config import MRR_TO_ARR, TOP_CUSTOMERS_SHEET
from ..source import (
    SourceContext,
    _is_customer_label_row,
    _parse_header_date,
    _to_float,
    dec_label,
    find_table_bounds,
    year_end_column_index,
)
from .top_sheet_common import (
    DATA_START,
    SN,
    TOP_N,
    add_conditional_formatting,
    apply_fmt,
    col_formats,
    col_layout,
    section_row_layout,
    set_column_widths,
    write_headers,
    write_top_n_customer_rows,
    write_other_row,
    write_sum_row,
    write_total_customers_row,
)


@dataclass
class _Customer:
    name: str
    src_row: int
    last_arr: float = 0.0
    since: str = ""


def _read_customers(ctx: SourceContext) -> list[_Customer]:
    ws = ctx.ws_in
    header_row = ctx.header_row
    cust_col = ctx.cust_col
    y2c = year_end_column_index(header_row, ws)
    last_year = max(ctx.years)
    last_col = y2c[last_year]

    _, _, first_dc = find_table_bounds(ws)
    date_cols: list[tuple[int, date]] = []
    for c in range(first_dc, ws.max_column + 1):
        d = _parse_header_date(ws.cell(header_row, c).value)
        if d is not None:
            date_cols.append((c, d))
    date_cols.sort(key=lambda t: t[1])

    customers: list[_Customer] = []
    for r in range(ctx.data_start_row, ctx.data_end_row + 1):
        label = ws.cell(r, cust_col).value
        if not _is_customer_label_row(label):
            continue
        cust = _Customer(
            name=str(label).strip(),
            src_row=r,
            last_arr=_to_float(ws.cell(r, last_col).value) * MRR_TO_ARR,
        )
        for col_idx, d in date_cols:
            if _to_float(ws.cell(r, col_idx).value) > 0:
                cust.since = f"{d:%b}-{d:%y}"
                break
        customers.append(cust)
    return customers


class TopCustomersTab:
    title = TOP_CUSTOMERS_SHEET

    def build(self, wb: Workbook, ctx: SourceContext) -> None:
        years = ctx.years
        n = len(years)
        layout = col_layout(n)
        total_cols = layout["since"]

        ic = ctx.intermediate
        src_cols = {y: get_column_letter(c) for y, c in ic.year_cols.items()}
        cust_cl = "A"
        rank_cl = get_column_letter(ic.rank_col)
        ds = ic.cust_start_row
        de = ic.cust_end_row

        customers = _read_customers(ctx)
        ranked = sorted(customers, key=lambda c: c.last_arr, reverse=True)
        ranked = [c for c in ranked if c.last_arr > 10]
        top10 = ranked[:TOP_N]
        num_top = len(top10)
        last_src_col = src_cols[max(years)]
        other_label = (
            f'="Other ("&COUNTIFS({SN}!${last_src_col}${ds}:'
            f'${last_src_col}${de},">"&10)'
            f'-{num_top}&" Customers)"'
        )

        rows = section_row_layout(DATA_START)
        top10_end = rows["data_end"]
        top10_total_row = rows["top10_total_row"]
        other_row = rows["other_row"]
        total_row = rows["total_row"]
        top5_row = rows["top5_row"]
        top10_row2 = rows["top10_row2"]

        ws = wb.create_sheet()
        ws.title = self.title[:31]

        last_lbl = dec_label(years[-1])
        write_headers(
            ws,
            years,
            layout,
            total_cols,
            title_a1=f"Top Customers as of {last_lbl}",
        )
        fmts = col_formats(layout, n)

        write_top_n_customer_rows(
            ws,
            DATA_START,
            top10,
            years,
            src_cols,
            layout,
            total_row,
            cust_cl,
            rank_cl,
            ds,
            de,
            fmts,
            total_cols,
        )

        write_sum_row(
            ws,
            top10_total_row,
            "Total Top 10",
            DATA_START,
            top10_end,
            years,
            layout,
            total_row,
            bold=True,
            border=True,
            total_cols=total_cols,
        )
        apply_fmt(ws, top10_total_row, fmts)

        write_other_row(
            ws,
            other_row,
            other_label,
            top10_total_row,
            total_row,
            years,
            layout,
            total_cols,
        )
        apply_fmt(ws, other_row, fmts)

        write_total_customers_row(
            ws,
            total_row,
            years,
            src_cols,
            layout,
            ds,
            de,
            total_cols,
        )
        apply_fmt(ws, total_row, fmts)

        top5_end = rows["top5_end"]
        write_sum_row(
            ws,
            top5_row,
            "Total Top 5",
            DATA_START,
            top5_end,
            years,
            layout,
            total_row,
            total_cols=total_cols,
        )
        apply_fmt(ws, top5_row, fmts)

        write_sum_row(
            ws,
            top10_row2,
            "Total Top 10",
            DATA_START,
            top10_end,
            years,
            layout,
            total_row,
            total_cols=total_cols,
        )
        apply_fmt(ws, top10_row2, fmts)

        add_conditional_formatting(ws, layout, n, DATA_START, top10_row2)

        from .cohort_customers import SECTION_GAP, append_cohort_sections_to_top_customers_sheet

        append_cohort_sections_to_top_customers_sheet(
            ws, ctx, top10_row2 + SECTION_GAP
        )

        set_column_widths(ws, total_cols)
