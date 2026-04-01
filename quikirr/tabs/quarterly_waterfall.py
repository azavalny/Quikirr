from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..config import MRR_TO_ARR, ORIGINAL_SHEET, QUARTERLY_SHEET
from ..source import SourceContext, quarter_label
from ..writer import write_summary_sheet


def build_quarterly_formula_values(
    quarter_keys: list[tuple[int, int]],
    quarter_col_letters: dict[tuple[int, int], str],
    data_start_row: int,
    data_end_row: int,
    src_sheet: str = ORIGINAL_SHEET,
    intermediate=None,
) -> tuple[list[str], list[dict[str, Any]]]:
    col_keys = [quarter_label(y, m) for y, m in quarter_keys]
    n = len(quarter_keys)
    sn = f"'{src_sheet}'"

    def _rng(key: tuple[int, int]) -> str:
        cl = quarter_col_letters[key]
        return f"{sn}!${cl}${data_start_row}:${cl}${data_end_row}"

    if intermediate:
        ic_sn = f"'{intermediate.sheet_name}'"

        def _ic_ref(metric_key: str, qk: tuple[int, int]) -> str:
            cl = get_column_letter(intermediate.quarter_cols[qk])
            row = intermediate.quarterly_rows[metric_key]
            return f"={ic_sn}!${cl}${row}"

    def _oc(i: int) -> str:
        return get_column_letter(i + 2)

    rows: list[dict[str, Any]] = []

    def new_row() -> dict[str, Any]:
        return {k: None for k in col_keys}

    D = "HATCH"
    cur = 3

    # ---- Quarterly Waterfall ----

    R_BOP = cur; cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        elif intermediate:
            r[col_keys[i]] = _ic_ref("bop", quarter_keys[i])
        else:
            r[col_keys[i]] = f"=SUM({_rng(quarter_keys[i-1])})*{MRR_TO_ARR}"
    rows.append({"label": "BoP ARR", "cells": r})

    R_NEW = cur; cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        elif intermediate:
            r[col_keys[i]] = _ic_ref("new", quarter_keys[i])
        else:
            pr, cr = _rng(quarter_keys[i - 1]), _rng(quarter_keys[i])
            r[col_keys[i]] = f"=SUMPRODUCT(({pr}<=0)*({cr}>0)*{cr})*{MRR_TO_ARR}"
    rows.append({"label": "New", "cells": r})

    R_UP = cur; cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        elif intermediate:
            r[col_keys[i]] = _ic_ref("up", quarter_keys[i])
        else:
            pr, cr = _rng(quarter_keys[i - 1]), _rng(quarter_keys[i])
            r[col_keys[i]] = (
                f"=SUMPRODUCT(({pr}>0)*({cr}>0)*({cr}>{pr})"
                f"*({cr}-{pr}))*{MRR_TO_ARR}"
            )
    rows.append({"label": "Upsell", "cells": r})

    R_DS = cur; cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        elif intermediate:
            r[col_keys[i]] = _ic_ref("ds", quarter_keys[i])
        else:
            pr, cr = _rng(quarter_keys[i - 1]), _rng(quarter_keys[i])
            r[col_keys[i]] = (
                f"=SUMPRODUCT(({pr}>0)*({cr}>0)*({cr}<{pr})"
                f"*({cr}-{pr}))*{MRR_TO_ARR}"
            )
    rows.append({"label": "(Downsell)", "cells": r})

    R_CH = cur; cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        elif intermediate:
            r[col_keys[i]] = _ic_ref("ch", quarter_keys[i])
        else:
            pr, cr = _rng(quarter_keys[i - 1]), _rng(quarter_keys[i])
            r[col_keys[i]] = f"=-SUMPRODUCT(({pr}>0)*({cr}<=0)*{pr})*{MRR_TO_ARR}"
    rows.append({"label": "(Churn)", "cells": r})

    R_EOP = cur; cur += 1
    r = new_row()
    for i in range(n):
        if intermediate:
            r[col_keys[i]] = _ic_ref("eop", quarter_keys[i])
        else:
            r[col_keys[i]] = f"=SUM({_rng(quarter_keys[i])})*{MRR_TO_ARR}"
    rows.append({"label": "EoP ARR", "cells": r, "eop": True})

    # ---- Growth metrics ----

    cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        else:
            oc = _oc(i)
            r[col_keys[i]] = f'=IF({oc}{R_BOP}=0,0,({oc}{R_EOP}/{oc}{R_BOP})^4-1)'
    rows.append({"label": "% Growth QoQ Annualized", "cells": r, "italic": True, "pct": True})

    cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        else:
            oc = _oc(i)
            r[col_keys[i]] = f'=IF({oc}{R_BOP}=0,0,{oc}{R_EOP}/{oc}{R_BOP}-1)'
    rows.append({"label": "% Growth QoQ", "cells": r, "italic": True, "pct": True})

    cur += 1
    r = new_row()
    for i in range(n):
        if i < 4:
            r[col_keys[i]] = D
        else:
            oc = _oc(i)
            oc4 = _oc(i - 4)
            r[col_keys[i]] = f'=IF({oc4}{R_EOP}=0,0,{oc}{R_EOP}/{oc4}{R_EOP}-1)'
    rows.append({"label": "% Growth YoY", "cells": r, "italic": True, "pct": True})

    cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        else:
            oc = _oc(i)
            r[col_keys[i]] = f"={oc}{R_EOP}-{oc}{R_BOP}"
    rows.append({"label": "Net New ARR QoQ", "cells": r, "italic": True, "currency": True})

    cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        else:
            oc = _oc(i)
            r[col_keys[i]] = f"={oc}{R_NEW}+{oc}{R_UP}"
    rows.append({"label": "New + Upsell QoQ", "cells": r, "italic": True, "currency": True})

    cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        else:
            oc = _oc(i)
            r[col_keys[i]] = f"={oc}{R_DS}+{oc}{R_CH}"
    rows.append({"label": "Downsell + Churn QoQ", "cells": r, "italic": True, "currency": True})

    # ---- Retention (QoQ Annualized) ----

    cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        else:
            oc = _oc(i)
            r[col_keys[i]] = (
                f'=IF({oc}{R_BOP}=0,0,'
                f'(({oc}{R_BOP}+{oc}{R_DS}+{oc}{R_CH})/{oc}{R_BOP})^4)'
            )
    rows.append({"label": "Gross Retention (QoQ Annualized)", "cells": r, "beige": True, "pct": True})

    cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        else:
            oc = _oc(i)
            r[col_keys[i]] = (
                f'=IF({oc}{R_BOP}=0,0,'
                f'(({oc}{R_BOP}+{oc}{R_CH})/{oc}{R_BOP})^4)'
            )
    rows.append({"label": "Loss-Only Retention (QoQ Annualized)", "cells": r, "beige": True, "pct": True})

    cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        else:
            oc = _oc(i)
            r[col_keys[i]] = (
                f'=IF({oc}{R_BOP}=0,0,'
                f'(({oc}{R_BOP}+{oc}{R_UP}+{oc}{R_DS}+{oc}{R_CH})/{oc}{R_BOP})^4)'
            )
    rows.append({"label": "Net Retention (QoQ Annualized)", "cells": r, "beige": True, "pct": True})

    # ---- Retention (LTM) ----
    # (BoP[i-3] + SUM of losses over 4 quarters) / BoP[i-3]

    cur += 1
    r = new_row()
    for i in range(n):
        if i < 4:
            r[col_keys[i]] = D
        else:
            lc, rc = _oc(i - 3), _oc(i)
            bop = f"{lc}{R_BOP}"
            r[col_keys[i]] = (
                f'=IF({bop}=0,0,'
                f"({bop}+SUM({lc}{R_DS}:{rc}{R_DS})"
                f"+SUM({lc}{R_CH}:{rc}{R_CH}))/{bop})"
            )
    rows.append({"label": "Gross Retention (LTM)", "cells": r, "beige": True, "pct": True})

    cur += 1
    r = new_row()
    for i in range(n):
        if i < 4:
            r[col_keys[i]] = D
        else:
            lc, rc = _oc(i - 3), _oc(i)
            bop = f"{lc}{R_BOP}"
            r[col_keys[i]] = (
                f'=IF({bop}=0,0,'
                f"({bop}+SUM({lc}{R_CH}:{rc}{R_CH}))/{bop})"
            )
    rows.append({"label": "Loss-Only Retention (LTM)", "cells": r, "beige": True, "pct": True})

    cur += 1
    r = new_row()
    for i in range(n):
        if i < 4:
            r[col_keys[i]] = D
        else:
            lc, rc = _oc(i - 3), _oc(i)
            bop = f"{lc}{R_BOP}"
            r[col_keys[i]] = (
                f'=IF({bop}=0,0,'
                f"({bop}+SUM({lc}{R_UP}:{rc}{R_UP})"
                f"+SUM({lc}{R_DS}:{rc}{R_DS})"
                f"+SUM({lc}{R_CH}:{rc}{R_CH}))/{bop})"
            )
    rows.append({"label": "Net Retention (LTM)", "cells": r, "beige": True, "pct": True})

    # ---- Customer Waterfall ----

    rows.append({"label": "__CW_HEADER__", "cells": {}})
    cur += 1

    R_BOPC = cur; cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        elif intermediate:
            r[col_keys[i]] = _ic_ref("bopc", quarter_keys[i])
        else:
            r[col_keys[i]] = f'=COUNTIF({_rng(quarter_keys[i-1])},">0")'
    rows.append({"label": "BoP Customers", "cells": r, "int": True})

    R_NEWC = cur; cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        elif intermediate:
            r[col_keys[i]] = _ic_ref("newc", quarter_keys[i])
        else:
            pr, cr = _rng(quarter_keys[i - 1]), _rng(quarter_keys[i])
            r[col_keys[i]] = f"=SUMPRODUCT(({pr}<=0)*({cr}>0)*1)"
    rows.append({"label": "New", "cells": r, "int": True})

    R_CHC = cur; cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        elif intermediate:
            r[col_keys[i]] = _ic_ref("chc", quarter_keys[i])
        else:
            pr, cr = _rng(quarter_keys[i - 1]), _rng(quarter_keys[i])
            r[col_keys[i]] = f"=SUMPRODUCT(({pr}>0)*({cr}<=0)*1)"
    rows.append({"label": "(Churn)", "cells": r, "int": True, "churn_count": True})

    R_EOPC = cur; cur += 1
    r = new_row()
    for i in range(n):
        if intermediate:
            r[col_keys[i]] = _ic_ref("eopc", quarter_keys[i])
        else:
            r[col_keys[i]] = f'=COUNTIF({_rng(quarter_keys[i])},">0")'
    rows.append({"label": "EoP Customers", "cells": r, "eop": True, "int": True})

    # ---- Customer derived metrics ----

    cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        else:
            oc = _oc(i)
            r[col_keys[i]] = f"={oc}{R_NEWC}-{oc}{R_CHC}"
    rows.append({"label": "Net Logos Added", "cells": r, "italic": True, "int": True})

    cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        else:
            oc = _oc(i)
            r[col_keys[i]] = f'=IF({oc}{R_BOPC}=0,0,{oc}{R_EOPC}/{oc}{R_BOPC}-1)'
    rows.append({"label": "Logo % Growth", "cells": r, "italic": True, "pct_or_dash": True, "pct": True})

    cur += 1
    r = new_row()
    for i in range(n):
        if i < 4:
            r[col_keys[i]] = D
        else:
            oc = _oc(i)
            oc4 = _oc(i - 4)
            r[col_keys[i]] = f'=IF({oc4}{R_EOPC}=0,0,{oc}{R_EOPC}/{oc4}{R_EOPC}-1)'
    rows.append({"label": "Logo % Growth (YoY)", "cells": r, "italic": True, "pct_or_dash": True, "pct": True})

    # ---- Logo retention ----

    cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        else:
            oc = _oc(i)
            r[col_keys[i]] = (
                f'=IF({oc}{R_BOPC}=0,0,'
                f'(({oc}{R_BOPC}-{oc}{R_CHC})/{oc}{R_BOPC})^4)'
            )
    rows.append({"label": "Gross Logo Retention (QoQ Annualized)", "cells": r, "beige": True, "pct": True})

    cur += 1
    r = new_row()
    for i in range(n):
        if i < 4:
            r[col_keys[i]] = D
        else:
            lc, rc = _oc(i - 3), _oc(i)
            bop = f"{lc}{R_BOPC}"
            r[col_keys[i]] = (
                f'=IF({bop}=0,0,'
                f"({bop}-SUM({lc}{R_CHC}:{rc}{R_CHC}))/{bop})"
            )
    rows.append({"label": "Gross Logo Retention (YoY)", "cells": r, "beige": True, "pct": True})

    # ---- Avg logo sizes ----

    R_AVGL = cur; cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        else:
            oc = _oc(i)
            r[col_keys[i]] = f'=IF({oc}{R_EOPC}=0,0,{oc}{R_EOP}/{oc}{R_EOPC})'
    rows.append({"label": "Avg. Logo Size ($ Actuals)", "cells": r, "currency": True})

    cur += 1
    r = new_row()
    for i in range(n):
        if i < 2:
            r[col_keys[i]] = D
        else:
            oc, poc = _oc(i), _oc(i - 1)
            r[col_keys[i]] = (
                f'=IF({poc}{R_AVGL}=0,0,'
                f'{oc}{R_AVGL}/{poc}{R_AVGL}-1)'
            )
    rows.append({"label": "Avg Logo Size QoQ Growth", "cells": r, "italic": True, "pct": True, "pct_parens_neg": True})

    R_AVGNL = cur; cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        else:
            oc = _oc(i)
            r[col_keys[i]] = f'=IF(OR({oc}{R_NEWC}=0,{oc}{R_NEW}=0),0,{oc}{R_NEW}/{oc}{R_NEWC})'
    rows.append({"label": "Avg. New Logo Size ($ Actuals)", "cells": r, "currency": True})

    cur += 1
    r = new_row()
    for i in range(n):
        if i < 2:
            r[col_keys[i]] = D
        else:
            oc, poc = _oc(i), _oc(i - 1)
            r[col_keys[i]] = (
                f'=IF({poc}{R_AVGNL}=0,0,'
                f'{oc}{R_AVGNL}/{poc}{R_AVGNL}-1)'
            )
    rows.append({"label": "Avg New Logo Size QoQ Growth", "cells": r, "italic": True, "pct": True, "pct_parens_neg": True})

    R_AVGCL = cur; cur += 1
    r = new_row()
    for i in range(n):
        if i == 0:
            r[col_keys[i]] = D
        else:
            oc = _oc(i)
            r[col_keys[i]] = f'=IF(OR({oc}{R_CHC}=0,{oc}{R_CH}=0),0,-{oc}{R_CH}/{oc}{R_CHC})'
    rows.append({"label": "Avg. Churned Logo Size ($ Actuals)", "cells": r, "currency": True})

    cur += 1
    r = new_row()
    for i in range(n):
        if i < 2:
            r[col_keys[i]] = D
        else:
            oc, poc = _oc(i), _oc(i - 1)
            r[col_keys[i]] = (
                f'=IF({poc}{R_AVGCL}=0,0,'
                f'{oc}{R_AVGCL}/{poc}{R_AVGCL}-1)'
            )
    rows.append({"label": "Avg Churned Logo Size QoQ Growth", "cells": r, "italic": True, "pct": True, "pct_parens_neg": True})

    return col_keys, rows


class QuarterlyWaterfallTab:
    title = QUARTERLY_SHEET

    def build(self, wb: Workbook, ctx: SourceContext) -> None:
        col_keys, row_specs = build_quarterly_formula_values(
            ctx.quarter_keys,
            ctx.quarter_col_letters,
            ctx.data_start_row,
            ctx.data_end_row,
            intermediate=ctx.intermediate,
        )
        ws = wb.create_sheet()
        ws.title = self.title[:31]
        write_summary_sheet(ws, self.title, col_keys, row_specs)
