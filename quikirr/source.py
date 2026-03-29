from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

from openpyxl.utils import get_column_letter

from .config import MRR_TO_ARR


def _is_excel_error_value(v: Any) -> bool:
    if isinstance(v, str):
        t = v.strip()
        return bool(t.startswith("#") and t.endswith("!"))
    return False


def _is_customer_label_row(label: Any) -> bool:
    if label is None:
        return False
    if isinstance(label, str):
        s = label.strip()
        if not s or _is_excel_error_value(s):
            return False
        if s in "-–—":
            return False
        return True
    return isinstance(label, (int, float))


def _to_float(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if _is_excel_error_value(v):
        return 0.0
    if isinstance(v, str) and v.strip() in "-–—":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _parse_header_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def find_table_bounds(ws) -> tuple[int, int, int]:
    """Return (header_row, customer_col, first_date_col)."""
    for r in range(1, min(ws.max_row, 50) + 1):
        date_cols: list[int] = []
        cust_col: int | None = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and re.search(r"customer", v, re.I):
                cust_col = c
            if _parse_header_date(v) is not None:
                date_cols.append(c)
        if len(date_cols) >= 3 and cust_col is not None:
            if date_cols[0] > cust_col:
                return r, cust_col, date_cols[0]
        if len(date_cols) >= 3:
            first_d = min(date_cols)
            cc = cust_col if cust_col is not None and cust_col < first_d else first_d - 1
            if cc >= 1:
                return r, cc, first_d
    raise ValueError("Could not locate header row with dates and customer column.")


def year_end_column_index(header_row: int, ws) -> dict[int, int]:
    """Map calendar year -> 1-based column index (last month in year, prefer Dec)."""
    _, _, first_dc = find_table_bounds(ws)
    by_year: dict[int, list[tuple[int, date]]] = {}
    for c in range(first_dc, ws.max_column + 1):
        d = _parse_header_date(ws.cell(header_row, c).value)
        if d is None:
            continue
        by_year.setdefault(d.year, []).append((c, d))
    chosen: dict[int, int] = {}
    for y, lst in by_year.items():
        dec = [t for t in lst if t[1].month == 12 and t[1].day == 31]
        if dec:
            chosen[y] = max(dec, key=lambda t: t[1])[0]
        else:
            chosen[y] = max(lst, key=lambda t: t[1])[0]
    return chosen


def read_customer_vectors(ws, header_row: int, cust_col: int, date_col: int) -> list[float]:
    vecs: list[float] = []
    for r in range(header_row + 1, ws.max_row + 1):
        label = ws.cell(r, cust_col).value
        if not _is_customer_label_row(label):
            continue
        vecs.append(_to_float(ws.cell(r, date_col).value) * MRR_TO_ARR)
    return vecs


def collect_snapshots(ws) -> tuple[int, int, dict[int, list[float]]]:
    header_row, cust_col, _ = find_table_bounds(ws)
    y2c = year_end_column_index(header_row, ws)
    years = sorted(y2c.keys())
    snaps: dict[int, list[float]] = {}
    for y in years:
        snaps[y] = read_customer_vectors(ws, header_row, cust_col, y2c[y])
    return header_row, cust_col, snaps


def find_data_bounds(ws, header_row: int, cust_col: int) -> tuple[int, int]:
    first: int | None = None
    last: int | None = None
    for r in range(header_row + 1, ws.max_row + 1):
        if _is_customer_label_row(ws.cell(r, cust_col).value):
            if first is None:
                first = r
            last = r
    if first is None:
        raise ValueError("No customer data rows found.")
    return first, last


def dec_label(y: int) -> str:
    return f"Dec-{str(y)[-2:]}"


_QUARTER_MONTH_ABBR = {3: "Mar", 6: "Jun", 9: "Sep", 12: "Dec"}


def quarter_label(year: int, month: int) -> str:
    return f"{_QUARTER_MONTH_ABBR[month]}-{str(year)[-2:]}"


def quarter_end_column_index(header_row: int, ws) -> dict[tuple[int, int], int]:
    """Map (year, quarter_end_month) -> 1-based column index."""
    _, _, first_dc = find_table_bounds(ws)
    by_quarter: dict[tuple[int, int], list[tuple[int, date]]] = {}
    for c in range(first_dc, ws.max_column + 1):
        d = _parse_header_date(ws.cell(header_row, c).value)
        if d is None:
            continue
        qm = ((d.month - 1) // 3 + 1) * 3
        key = (d.year, qm)
        by_quarter.setdefault(key, []).append((c, d))
    chosen: dict[tuple[int, int], int] = {}
    for key, lst in by_quarter.items():
        _, qm = key
        exact = [t for t in lst if t[1].month == qm]
        if exact:
            chosen[key] = max(exact, key=lambda t: t[1])[0]
        else:
            chosen[key] = max(lst, key=lambda t: t[1])[0]
    return chosen


@dataclass
class SourceContext:
    """Everything a tab needs to build its sheet."""
    ws_in: Any
    header_row: int
    cust_col: int
    years: list[int]
    year_col_letters: dict[int, str]
    data_start_row: int
    data_end_row: int
    snaps: dict[int, list[float]] = field(default_factory=dict)
    quarter_keys: list[tuple[int, int]] = field(default_factory=list)
    quarter_col_letters: dict[tuple[int, int], str] = field(default_factory=dict)

    @classmethod
    def from_worksheet(cls, ws) -> SourceContext:
        header_row, cust_col, snaps = collect_snapshots(ws)
        years = sorted(snaps.keys())
        if len(years) < 1:
            raise ValueError("No year-end columns found.")
        y2c = year_end_column_index(header_row, ws)
        year_col_ltrs = {y: get_column_letter(c) for y, c in y2c.items()}
        data_start, data_end = find_data_bounds(ws, header_row, cust_col)
        q2c = quarter_end_column_index(header_row, ws)
        q_keys = sorted(q2c.keys())
        q_col_ltrs = {k: get_column_letter(c) for k, c in q2c.items()}
        return cls(
            ws_in=ws,
            header_row=header_row,
            cust_col=cust_col,
            years=years,
            year_col_letters=year_col_ltrs,
            data_start_row=data_start,
            data_end_row=data_end,
            snaps=snaps,
            quarter_keys=q_keys,
            quarter_col_letters=q_col_ltrs,
        )
