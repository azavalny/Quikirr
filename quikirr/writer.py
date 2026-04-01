from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Border
from openpyxl.utils import get_column_letter

from .source import _is_excel_error_value
from .styles import (
    ACCOUNTING_FMT,
    CHURN_COUNT_FMT,
    FILL_BEIGE,
    FILL_BLUE,
    FILL_HATCH,
    FONT_BOLD,
    FONT_ITALIC,
    FONT_WHITE_BOLD,
    INT_FMT,
    PCT_FMT,
    PCT_PARENS_NEG_FMT,
    THIN,
)


def _sanitize_copy_value(v: Any) -> Any:
    if _is_excel_error_value(v):
        return None
    return v


def _display_currency(val: Any) -> Any:
    if val is None or isinstance(val, str):
        return val
    if isinstance(val, (int, float)):
        return int(round(float(val)))
    return val


def _display_pct(val: Any) -> Any:
    if val is None or isinstance(val, str):
        return val
    if isinstance(val, (int, float)):
        return round(float(val), 8)
    return val


def copy_original_sheet(src_ws, dst_ws) -> None:
    for r in range(1, src_ws.max_row + 1):
        for c in range(1, src_ws.max_column + 1):
            dst_ws.cell(r, c, value=_sanitize_copy_value(src_ws.cell(r, c).value))


def write_summary_sheet(
    ws,
    title: str,
    col_keys: list[str],
    row_specs: list[dict[str, Any]],
) -> None:
    ncols = len(col_keys)
    last_col_letter = get_column_letter(1 + ncols)

    c1 = ws["A1"]
    c1.value = title
    c1.fill = FILL_BLUE
    c1.font = FONT_WHITE_BOLD
    c1.alignment = Alignment(horizontal="center")
    for j in range(2, 2 + ncols):
        ws.cell(1, j).fill = FILL_BLUE

    for j, ck in enumerate(col_keys, start=2):
        cell = ws.cell(2, j, value=ck)
        cell.fill = FILL_BLUE
        cell.font = FONT_WHITE_BOLD
        cell.alignment = Alignment(horizontal="right")
    ws.cell(2, 1).fill = FILL_BLUE
    ws.cell(2, 1).value = None

    current_row = 3

    for spec in row_specs:
        if spec.get("label") == "__CW_HEADER__":
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=1 + ncols)
            h = ws.cell(current_row, 1, value="Customer Waterfall")
            h.fill = FILL_BLUE
            h.font = FONT_WHITE_BOLD
            for j in range(2, 2 + ncols):
                ws.cell(current_row, j).fill = FILL_BLUE
            current_row += 1
            continue

        label = spec["label"]
        cells = spec["cells"]
        pct_fmt_row = PCT_PARENS_NEG_FMT if spec.get("pct_parens_neg") else PCT_FMT
        ws.cell(current_row, 1, value=label)
        ws.cell(current_row, 1).alignment = Alignment(horizontal="left")

        is_italic_label = spec.get("italic")
        if is_italic_label:
            ws.cell(current_row, 1).font = FONT_ITALIC

        for j, ck in enumerate(col_keys, start=2):
            cell = ws.cell(current_row, j)
            val = cells.get(ck)
            if val == "HATCH":
                cell.value = None
                cell.fill = FILL_HATCH
            elif val is None:
                cell.value = None
            elif spec.get("pct_or_dash") and isinstance(val, (int, float)):
                cell.value = _display_pct(val)
                cell.number_format = pct_fmt_row
                cell.alignment = Alignment(horizontal="right")
                if is_italic_label:
                    cell.font = FONT_ITALIC
            elif spec.get("pct"):
                cell.value = _display_pct(val)
                cell.number_format = pct_fmt_row
                cell.alignment = Alignment(horizontal="right")
                if is_italic_label:
                    cell.font = FONT_ITALIC
            elif spec.get("currency"):
                cell.value = _display_currency(val)
                cell.number_format = ACCOUNTING_FMT
                cell.alignment = Alignment(horizontal="right")
                if is_italic_label:
                    cell.font = FONT_ITALIC
            elif spec.get("int"):
                cell.value = val
                cell.alignment = Alignment(horizontal="right")
                if spec.get("churn_count"):
                    cell.number_format = CHURN_COUNT_FMT
                else:
                    cell.number_format = INT_FMT
                if is_italic_label:
                    cell.font = FONT_ITALIC
            else:
                cell.value = _display_currency(val)
                cell.number_format = ACCOUNTING_FMT
                cell.alignment = Alignment(horizontal="right")

            if spec.get("beige"):
                cell.fill = FILL_BEIGE
                ws.cell(current_row, 1).fill = FILL_BEIGE
                ws.cell(current_row, 1).font = FONT_BOLD
                cell.font = FONT_BOLD
                if isinstance(cell.value, (int, float)):
                    cell.number_format = pct_fmt_row

            if spec.get("eop"):
                cell.font = FONT_BOLD
                ws.cell(current_row, 1).font = FONT_BOLD
                cell.border = Border(top=THIN, bottom=THIN)
                ws.cell(current_row, 1).border = Border(top=THIN, bottom=THIN)
                if not spec.get("int"):
                    cell.number_format = ACCOUNTING_FMT

        current_row += 1

    ws.column_dimensions["A"].width = 36
    for j in range(2, 2 + ncols):
        ws.column_dimensions[get_column_letter(j)].width = 14
