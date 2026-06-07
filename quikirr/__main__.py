from __future__ import annotations

import argparse
import io
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .config import (
    DEFAULT_OUTPUT,
    DEFAULT_SOURCE,
    INTERMEDIATE_SHEET,
    OUTPUT_SHEET,
    ORIGINAL_SHEET,
    QUARTERLY_SHEET,
    SOURCE_SHEET,
    TOP_CUSTOMERS_SHEET,
)
from .source import SourceContext, find_table_bounds
from .tabs import TABS
from .verify import assert_bridge

_WORKBOOK_TAB_ORDER = (
    OUTPUT_SHEET,
    QUARTERLY_SHEET,
    TOP_CUSTOMERS_SHEET,
    INTERMEDIATE_SHEET,
    ORIGINAL_SHEET,
)


def _reorder_workbook_sheets(wb, titles: tuple[str, ...]) -> None:
    by_title = {ws.title: ws for ws in wb.worksheets}
    ordered = [by_title[t] for t in titles]
    extra = [ws for ws in wb.worksheets if ws.title not in titles]
    wb._sheets = ordered + extra


def _resolve_source_worksheet(wb_in):
    if SOURCE_SHEET in wb_in.sheetnames:
        return wb_in[SOURCE_SHEET]
    by_norm = {name.strip().casefold(): name for name in wb_in.sheetnames}
    match = by_norm.get(SOURCE_SHEET.casefold())
    if match is not None:
        return wb_in[match]
    for name in wb_in.sheetnames:
        try:
            find_table_bounds(wb_in[name])
        except ValueError:
            continue
        return wb_in[name]
    raise KeyError(f"Missing sheet {SOURCE_SHEET!r}")


def _build_output_workbook(wb_in, verify: bool = True) -> Workbook:
    ctx = SourceContext.from_worksheet(_resolve_source_worksheet(wb_in))

    if verify:
        assert_bridge(ctx.years, ctx.snaps)

    wb_out = Workbook()
    for tab in TABS:
        tab.build(wb_out, ctx)

    _reorder_workbook_sheets(wb_out, _WORKBOOK_TAB_ORDER)
    return wb_out


def run(source: Path, output: Path, verify: bool = True) -> None:
    wb_in = load_workbook(source, data_only=True)
    wb_out = _build_output_workbook(wb_in, verify)
    wb_out.save(output)
    wb_in.close()


def run_to_bytes(source_bytes: bytes, verify: bool = False) -> bytes:
    wb_in = load_workbook(io.BytesIO(source_bytes), data_only=True)
    wb_out = _build_output_workbook(wb_in, verify)
    buf = io.BytesIO()
    wb_out.save(buf)
    wb_in.close()
    return buf.getvalue()


def main() -> None:
    p = argparse.ArgumentParser(description="Build Annual Waterfall Excel from Source.xlsx")
    p.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    p.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    p.add_argument("--no-verify", action="store_true")
    args = p.parse_args()
    run(args.source, args.output, verify=not args.no_verify)
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
