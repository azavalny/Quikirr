from __future__ import annotations

from openpyxl.styles import Font, PatternFill, Side

BLUE = "1F4E79"
WHITE = "FFFFFF"
BEIGE = "F2E4D8"
GREY_HATCH_FG = "D9D9D9"
GREY_HATCH_BG = "F2F2F2"

THIN = Side(style="thin", color="000000")

ACCOUNTING_FMT = '_("$"* #,##0_);_("$"* (#,##0);_("$"* "-"??_);_(@_)'
PCT_FMT = '0%_);(0%);" - "'
PCT_PARENS_NEG_FMT = '0%_);(0%);" - "'
CHURN_COUNT_FMT = '(#,##0);(#,##0);" - "'
INT_FMT = '_(* #,##0_);_(* (#,##0);_(* " - "?_);_(@_)'
MULT_FMT = '0.0"x"_);(0.0"x");" - "'

FILL_BLUE = PatternFill(fill_type="solid", fgColor=BLUE)
FILL_BEIGE = PatternFill(fill_type="solid", fgColor=BEIGE)
FILL_HATCH = PatternFill(patternType="lightDown", fgColor=GREY_HATCH_FG, bgColor=GREY_HATCH_BG)
FILL_GREEN_COND = PatternFill(fill_type="solid", fgColor="C6EFCE")
FILL_RED_COND = PatternFill(fill_type="solid", fgColor="FFC7CE")

FONT_WHITE_BOLD = Font(bold=True, color=WHITE)
FONT_BOLD = Font(bold=True)
FONT_ITALIC = Font(italic=True)
