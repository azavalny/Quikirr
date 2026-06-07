from __future__ import annotations

import calendar
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border
from openpyxl.utils import get_column_letter

from quikirr.styles import (
    ACCOUNTING_FMT,
    FILL_BLUE,
    FONT_BOLD,
    FONT_ITALIC,
    FONT_WHITE_BOLD,
    INT_FMT,
    MULT_FMT,
    PCT_FMT,
    THIN,
)
from quikirr.verify import compute_metrics_for_year

WORKBOOK = Path(__file__).resolve().parent / "Example Databook.xlsx"
SOURCE_SHEET = "Source Data"

YEARS = [2022, 2023, 2024, 2025]
HEADER_ROW = 3
CUST_COL = 2
FIRST_DATE_COL = 3
MRR_TO_ARR = 12

COMPANY = "Project Atlas"

DATA_COL0 = 2  # column B holds FY2022

COGS_PCT = {2022: 0.27, 2023: 0.25, 2024: 0.23, 2025: 0.21}
SM_PCT = {2022: 0.45, 2023: 0.42, 2024: 0.39, 2025: 0.37}
RD_PCT = {2022: 0.22, 2023: 0.21, 2024: 0.20, 2025: 0.19}
GA_PCT = {2022: 0.15, 2023: 0.14, 2024: 0.13, 2025: 0.12}
SERVICES_PCT = 0.08
DA_PCT = 0.03


def _year_end_col(year_index: int) -> int:
    return FIRST_DATE_COL + year_index * 12


def _year_month_cols(year_index: int) -> list[int]:
    if year_index == 0:
        return [FIRST_DATE_COL]
    start = FIRST_DATE_COL + (year_index - 1) * 12 + 1
    end = FIRST_DATE_COL + year_index * 12
    return list(range(start, end + 1))


def _col_to_year_index(col: int) -> int:
    if col == FIRST_DATE_COL:
        return 0
    return 1 + (col - FIRST_DATE_COL - 1) // 12


def _add_months(d: date, k: int) -> date:
    m = d.month - 1 + k
    y = d.year + m // 12
    m = m % 12 + 1
    return date(y, m, calendar.monthrange(y, m)[1])


def _normalize_source_data(ws, cust_rows: list[int]) -> None:
    """Rewrite formula-driven headers/IDs as literals so data_only readers
    (e.g. the email bot) can parse the Source Data tab."""
    last_date_col = _year_end_col(len(YEARS) - 1)
    anchor = ws.cell(HEADER_ROW, FIRST_DATE_COL).value
    if isinstance(anchor, datetime):
        anchor = anchor.date()
    if isinstance(anchor, date):
        for k, c in enumerate(range(FIRST_DATE_COL, last_date_col + 1)):
            d = _add_months(anchor, k)
            cell = ws.cell(HEADER_ROW, c)
            cell.value = datetime(d.year, d.month, d.day)
            cell.number_format = "m/d/yyyy"
    for i, r in enumerate(cust_rows, start=1):
        ws.cell(r, CUST_COL).value = i


def _num(v: object) -> float:
    return float(v) if isinstance(v, (int, float)) else 0.0


def _read_source(ws) -> dict:
    last_date_col = _year_end_col(len(YEARS) - 1)
    cust_rows: list[int] = []
    for r in range(4, ws.max_row + 1):
        if any(
            isinstance(ws.cell(r, c).value, (int, float))
            for c in range(FIRST_DATE_COL, last_date_col + 1)
        ):
            cust_rows.append(r)

    arr_by_year: dict[int, list[float]] = {}
    for i, y in enumerate(YEARS):
        col = _year_end_col(i)
        arr_by_year[y] = [
            _num(ws.cell(rr, col).value) * MRR_TO_ARR for rr in cust_rows
        ]

    avg_arr_by_year: dict[int, float] = {}
    for i, y in enumerate(YEARS):
        cols = _year_month_cols(i)
        total = 0.0
        for rr in cust_rows:
            vals = [_num(ws.cell(rr, c).value) for c in cols]
            total += (sum(vals) / len(vals)) * MRR_TO_ARR
        avg_arr_by_year[y] = total

    first_year_idx: list[int] = []
    for rr in cust_rows:
        fi = None
        for c in range(FIRST_DATE_COL, last_date_col + 1):
            if _num(ws.cell(rr, c).value) > 0:
                fi = _col_to_year_index(c)
                break
        first_year_idx.append(fi if fi is not None else -1)

    return dict(
        cust_rows=cust_rows,
        arr_by_year=arr_by_year,
        avg_arr_by_year=avg_arr_by_year,
        first_year_idx=first_year_idx,
    )


def _title(ws, text: str, last_col: int, subtitle: str | None = None) -> None:
    c = ws.cell(1, 1, value=text)
    c.fill = FILL_BLUE
    c.font = FONT_WHITE_BOLD
    c.alignment = Alignment(horizontal="left")
    for col in range(2, last_col + 1):
        ws.cell(1, col).fill = FILL_BLUE
    if subtitle is not None:
        s = ws.cell(2, 1, value=subtitle)
        s.font = FONT_ITALIC


def _year_headers(ws, row: int, first_data_col: int = DATA_COL0) -> None:
    for i, y in enumerate(YEARS):
        cell = ws.cell(row, first_data_col + i, value=f"FY{y}")
        cell.fill = FILL_BLUE
        cell.font = FONT_WHITE_BOLD
        cell.alignment = Alignment(horizontal="right")
    ws.cell(row, 1).fill = FILL_BLUE
    ws.cell(row, 1).font = FONT_WHITE_BOLD


def _set_widths(ws, last_col: int, label_w: int = 34) -> None:
    ws.column_dimensions["A"].width = label_w
    for j in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(j)].width = 15


def _label(ws, row: int, text: str, bold: bool = False, indent: int = 0) -> None:
    c = ws.cell(row, 1, value=text)
    if bold:
        c.font = FONT_BOLD
    if indent:
        c.alignment = Alignment(indent=indent)


def _apply_fmt(ws, row: int, fmt: str, first_col: int = DATA_COL0) -> None:
    for i in range(len(YEARS)):
        ws.cell(row, first_col + i).number_format = fmt


def _border_row(ws, row: int, last_col: int) -> None:
    for c in range(1, last_col + 1):
        ws.cell(row, c).border = Border(top=THIN, bottom=THIN)


def build_arr_bridge(wb, data: dict) -> str:
    name = "ARR Bridge"
    ws = wb.create_sheet(name)
    last_col = DATA_COL0 + len(YEARS) - 1
    _title(ws, f"{COMPANY} - ARR Bridge", last_col, "($ in actuals)")
    _year_headers(ws, HEADER_ROW)

    metrics = {}
    for i, y in enumerate(YEARS):
        prev = data["arr_by_year"][YEARS[i - 1]] if i > 0 else None
        metrics[y] = compute_metrics_for_year(prev, data["arr_by_year"][y])

    rows = {
        "bop": 4,
        "new": 5,
        "up": 6,
        "down": 7,
        "churn": 8,
        "eop": 9,
        "grr": 11,
        "nrr": 12,
        "growth": 13,
    }
    _label(ws, rows["bop"], "Beginning ARR")
    _label(ws, rows["new"], "(+) New", indent=1)
    _label(ws, rows["up"], "(+) Expansion", indent=1)
    _label(ws, rows["down"], "(-) Contraction", indent=1)
    _label(ws, rows["churn"], "(-) Churn", indent=1)
    _label(ws, rows["eop"], "Ending ARR", bold=True)
    _label(ws, rows["grr"], "Gross $ Retention")
    _label(ws, rows["nrr"], "Net $ Retention")
    _label(ws, rows["growth"], "YoY ARR Growth")

    for i, y in enumerate(YEARS):
        col = DATA_COL0 + i
        cl = get_column_letter(col)
        m = metrics[y]
        if m.needs_prior:
            ws.cell(rows["bop"], col).value = m.bop_arr
            ws.cell(rows["new"], col).value = m.new_arr
            ws.cell(rows["up"], col).value = m.upsell
            ws.cell(rows["down"], col).value = m.downsell
            ws.cell(rows["churn"], col).value = m.churn_arr
            ws.cell(rows["eop"], col).value = (
                f"={cl}{rows['bop']}+{cl}{rows['new']}+{cl}{rows['up']}"
                f"+{cl}{rows['down']}+{cl}{rows['churn']}"
            )
            ws.cell(rows["grr"], col).value = (
                f"=({cl}{rows['bop']}+{cl}{rows['down']}+{cl}{rows['churn']})"
                f"/{cl}{rows['bop']}"
            )
            ws.cell(rows["nrr"], col).value = (
                f"=({cl}{rows['bop']}+{cl}{rows['up']}+{cl}{rows['down']}"
                f"+{cl}{rows['churn']})/{cl}{rows['bop']}"
            )
            prev_cl = get_column_letter(col - 1)
            ws.cell(rows["growth"], col).value = (
                f"={cl}{rows['eop']}/{prev_cl}{rows['eop']}-1"
            )
        else:
            ws.cell(rows["eop"], col).value = m.eop_arr

    for key in ("bop", "new", "up", "down", "churn", "eop"):
        _apply_fmt(ws, rows[key], ACCOUNTING_FMT)
    for key in ("grr", "nrr", "growth"):
        _apply_fmt(ws, rows[key], PCT_FMT)
    _border_row(ws, rows["eop"], last_col)
    _set_widths(ws, last_col)
    ws.sheet_view.showGridLines = False
    return name


def build_cohort_retention(wb, data: dict) -> str:
    name = "Cohort Retention"
    ws = wb.create_sheet(name)
    last_col = 3 + len(YEARS)  # label + logos + initial + 4 years
    _title(ws, f"{COMPANY} - Net Dollar Retention by Cohort", last_col, "($ in actuals)")

    hdr = HEADER_ROW
    ws.cell(hdr, 1, value="Acquisition Cohort")
    ws.cell(hdr, 2, value="Logos")
    ws.cell(hdr, 3, value="Initial ARR")
    for i, y in enumerate(YEARS):
        ws.cell(hdr, 4 + i, value=f"FY{y}")
    for c in range(1, last_col + 1):
        ws.cell(hdr, c).fill = FILL_BLUE
        ws.cell(hdr, c).font = FONT_WHITE_BOLD
        ws.cell(hdr, c).alignment = Alignment(horizontal="right")
    ws.cell(hdr, 1).alignment = Alignment(horizontal="left")

    cust_rows = data["cust_rows"]
    fyi = data["first_year_idx"]
    arr_by_year = data["arr_by_year"]

    labels = {0: "FY2022 & Prior", 1: "FY2023", 2: "FY2024", 3: "FY2025"}
    for ci in range(len(YEARS)):
        row = hdr + 1 + ci
        members = [idx for idx, f in enumerate(fyi) if f == ci]
        cohort_year = YEARS[ci]
        initial = sum(arr_by_year[cohort_year][idx] for idx in members)
        ws.cell(row, 1, value=labels[ci])
        ws.cell(row, 2, value=len(members)).number_format = INT_FMT
        ws.cell(row, 3, value=initial).number_format = ACCOUNTING_FMT
        for yi, y in enumerate(YEARS):
            cell = ws.cell(row, 4 + yi)
            if yi < ci or initial <= 0:
                continue
            retained = sum(arr_by_year[y][idx] for idx in members)
            cell.value = retained / initial
            cell.number_format = PCT_FMT

    _set_widths(ws, last_col, label_w=22)
    ws.column_dimensions["C"].width = 15
    ws.sheet_view.showGridLines = False
    return name


def build_pnl(wb, data: dict) -> str:
    name = "P&L"
    ws = wb.create_sheet(name)
    last_col = DATA_COL0 + len(YEARS) - 1
    _title(ws, f"{COMPANY} - Income Statement", last_col, "($ in actuals)")
    _year_headers(ws, HEADER_ROW)

    rows = {
        "sub": 4,
        "svc": 5,
        "rev": 6,
        "cogs": 7,
        "gp": 8,
        "gm": 9,
        "sm": 11,
        "rd": 12,
        "ga": 13,
        "opex": 14,
        "ebitda": 15,
        "ebitda_m": 16,
        "da": 17,
        "ebit": 18,
    }
    _label(ws, rows["sub"], "Subscription revenue", indent=1)
    _label(ws, rows["svc"], "Services revenue", indent=1)
    _label(ws, rows["rev"], "Total Revenue", bold=True)
    _label(ws, rows["cogs"], "Cost of revenue", indent=1)
    _label(ws, rows["gp"], "Gross Profit", bold=True)
    _label(ws, rows["gm"], "Gross Margin %")
    _label(ws, 10, "Operating expenses", bold=True)
    _label(ws, rows["sm"], "Sales & marketing", indent=1)
    _label(ws, rows["rd"], "Research & development", indent=1)
    _label(ws, rows["ga"], "General & administrative", indent=1)
    _label(ws, rows["opex"], "Total Operating Expenses", bold=True)
    _label(ws, rows["ebitda"], "EBITDA", bold=True)
    _label(ws, rows["ebitda_m"], "EBITDA Margin %")
    _label(ws, rows["da"], "Depreciation & amortization", indent=1)
    _label(ws, rows["ebit"], "EBIT", bold=True)

    for i, y in enumerate(YEARS):
        col = DATA_COL0 + i
        cl = get_column_letter(col)
        sub = data["avg_arr_by_year"][y]
        ws.cell(rows["sub"], col).value = sub
        ws.cell(rows["svc"], col).value = sub * SERVICES_PCT
        ws.cell(rows["rev"], col).value = f"={cl}{rows['sub']}+{cl}{rows['svc']}"
        ws.cell(rows["cogs"], col).value = f"=-{cl}{rows['rev']}*{COGS_PCT[y]}"
        ws.cell(rows["gp"], col).value = f"={cl}{rows['rev']}+{cl}{rows['cogs']}"
        ws.cell(rows["gm"], col).value = f"={cl}{rows['gp']}/{cl}{rows['rev']}"
        ws.cell(rows["sm"], col).value = f"=-{cl}{rows['rev']}*{SM_PCT[y]}"
        ws.cell(rows["rd"], col).value = f"=-{cl}{rows['rev']}*{RD_PCT[y]}"
        ws.cell(rows["ga"], col).value = f"=-{cl}{rows['rev']}*{GA_PCT[y]}"
        ws.cell(rows["opex"], col).value = (
            f"={cl}{rows['sm']}+{cl}{rows['rd']}+{cl}{rows['ga']}"
        )
        ws.cell(rows["ebitda"], col).value = (
            f"={cl}{rows['gp']}+{cl}{rows['opex']}"
        )
        ws.cell(rows["ebitda_m"], col).value = (
            f"={cl}{rows['ebitda']}/{cl}{rows['rev']}"
        )
        ws.cell(rows["da"], col).value = f"=-{cl}{rows['rev']}*{DA_PCT}"
        ws.cell(rows["ebit"], col).value = f"={cl}{rows['ebitda']}+{cl}{rows['da']}"

    for key in ("sub", "svc", "rev", "cogs", "gp", "sm", "rd", "ga", "opex",
                "ebitda", "da", "ebit"):
        _apply_fmt(ws, rows[key], ACCOUNTING_FMT)
    for key in ("gm", "ebitda_m"):
        _apply_fmt(ws, rows[key], PCT_FMT)
    _border_row(ws, rows["rev"], last_col)
    _border_row(ws, rows["gp"], last_col)
    _border_row(ws, rows["ebitda"], last_col)
    _set_widths(ws, last_col)
    ws.sheet_view.showGridLines = False
    return name


def build_customer_unit_econ(wb, data: dict, bridge: str, pnl: str) -> str:
    name = "Customer & Unit Economics"
    ws = wb.create_sheet(name)
    last_col = DATA_COL0 + len(YEARS) - 1
    _title(ws, f"{COMPANY} - Customer & Unit Economics", last_col, "($ in actuals)")
    _year_headers(ws, HEADER_ROW)

    b = f"'{bridge}'"
    p = f"'{pnl}'"

    rows = {
        "bop_logo": 4,
        "new_logo": 5,
        "churn_logo": 6,
        "eop_logo": 7,
        "logo_churn": 8,
        "acv": 10,
        "grr": 11,
        "nrr": 12,
        "cac": 14,
        "payback": 15,
        "ltv": 16,
        "ltv_cac": 17,
        "magic": 18,
    }
    _label(ws, rows["bop_logo"], "Beginning customers")
    _label(ws, rows["new_logo"], "(+) New customers", indent=1)
    _label(ws, rows["churn_logo"], "(-) Churned customers", indent=1)
    _label(ws, rows["eop_logo"], "Ending customers", bold=True)
    _label(ws, rows["logo_churn"], "Logo churn %")
    _label(ws, rows["acv"], "Average ACV")
    _label(ws, rows["grr"], "Gross $ Retention")
    _label(ws, rows["nrr"], "Net $ Retention")
    _label(ws, rows["cac"], "CAC (per new logo)")
    _label(ws, rows["payback"], "CAC payback (months)")
    _label(ws, rows["ltv"], "LTV")
    _label(ws, rows["ltv_cac"], "LTV / CAC")
    _label(ws, rows["magic"], "Magic number")

    metrics = {}
    for i, y in enumerate(YEARS):
        prev = data["arr_by_year"][YEARS[i - 1]] if i > 0 else None
        metrics[y] = compute_metrics_for_year(prev, data["arr_by_year"][y])

    for i, y in enumerate(YEARS):
        col = DATA_COL0 + i
        cl = get_column_letter(col)
        m = metrics[y]
        if m.needs_prior:
            ws.cell(rows["bop_logo"], col).value = m.bop_cust
            ws.cell(rows["new_logo"], col).value = m.new_cust
            ws.cell(rows["churn_logo"], col).value = m.churn_cust
            ws.cell(rows["eop_logo"], col).value = (
                f"={cl}{rows['bop_logo']}+{cl}{rows['new_logo']}"
                f"-{cl}{rows['churn_logo']}"
            )
            ws.cell(rows["logo_churn"], col).value = (
                f"={cl}{rows['churn_logo']}/{cl}{rows['bop_logo']}"
            )
        else:
            ws.cell(rows["eop_logo"], col).value = m.eop_cust

        ws.cell(rows["acv"], col).value = (
            f"={b}!{cl}9/{cl}{rows['eop_logo']}"
        )
        ws.cell(rows["grr"], col).value = f"={b}!{cl}11"
        ws.cell(rows["nrr"], col).value = f"={b}!{cl}12"

        if m.needs_prior:
            ws.cell(rows["cac"], col).value = (
                f"=-{p}!{cl}11/{cl}{rows['new_logo']}"
            )
            ws.cell(rows["payback"], col).value = (
                f"={cl}{rows['cac']}/(({b}!{cl}5/{cl}{rows['new_logo']})"
                f"*{p}!{cl}9)*12"
            )
            ws.cell(rows["ltv"], col).value = (
                f"={cl}{rows['acv']}*{p}!{cl}9/{cl}{rows['logo_churn']}"
            )
            ws.cell(rows["ltv_cac"], col).value = (
                f"={cl}{rows['ltv']}/{cl}{rows['cac']}"
            )
            prev_cl = get_column_letter(col - 1)
            ws.cell(rows["magic"], col).value = (
                f"=({b}!{cl}9-{b}!{prev_cl}9)/-{p}!{prev_cl}11"
            )

    for key in ("bop_logo", "new_logo", "churn_logo", "eop_logo"):
        _apply_fmt(ws, rows[key], INT_FMT)
    for key in ("acv", "cac", "ltv"):
        _apply_fmt(ws, rows[key], ACCOUNTING_FMT)
    for key in ("logo_churn", "grr", "nrr"):
        _apply_fmt(ws, rows[key], PCT_FMT)
    _apply_fmt(ws, rows["payback"], '0.0" mo"')
    for key in ("ltv_cac", "magic"):
        _apply_fmt(ws, rows[key], MULT_FMT)
    _border_row(ws, rows["eop_logo"], last_col)
    _set_widths(ws, last_col)
    ws.sheet_view.showGridLines = False
    return name


def build_kpi_summary(wb, bridge: str, pnl: str, econ: str) -> str:
    name = "KPI Summary"
    ws = wb.create_sheet(name)
    last_col = DATA_COL0 + len(YEARS) - 1
    _title(ws, f"{COMPANY} - KPI Summary", last_col, "($ in actuals)")
    _year_headers(ws, HEADER_ROW)

    b = f"'{bridge}'"
    p = f"'{pnl}'"
    e = f"'{econ}'"

    specs = [
        ("Growth", None, None, True),
        ("Ending ARR", f"{b}!{{cl}}9", ACCOUNTING_FMT, False),
        ("YoY ARR growth %", f"{b}!{{cl}}13", PCT_FMT, False),
        ("Net $ retention", f"{b}!{{cl}}12", PCT_FMT, False),
        ("Gross $ retention", f"{b}!{{cl}}11", PCT_FMT, False),
        ("Customers & Logos", None, None, True),
        ("Ending customers", f"{e}!{{cl}}7", INT_FMT, False),
        ("New customers", f"{e}!{{cl}}5", INT_FMT, False),
        ("Churned customers", f"{e}!{{cl}}6", INT_FMT, False),
        ("Logo churn %", f"{e}!{{cl}}8", PCT_FMT, False),
        ("Average ACV", f"{e}!{{cl}}10", ACCOUNTING_FMT, False),
        ("Financials", None, None, True),
        ("Total revenue", f"{p}!{{cl}}6", ACCOUNTING_FMT, False),
        ("Gross margin %", f"{p}!{{cl}}9", PCT_FMT, False),
        ("EBITDA", f"{p}!{{cl}}15", ACCOUNTING_FMT, False),
        ("EBITDA margin %", f"{p}!{{cl}}16", PCT_FMT, False),
        ("Rule of 40", f"{b}!{{cl}}13+{p}!{{cl}}16", PCT_FMT, False),
        ("Unit Economics", None, None, True),
        ("CAC (per new logo)", f"{e}!{{cl}}14", ACCOUNTING_FMT, False),
        ("CAC payback (months)", f"{e}!{{cl}}15", '0.0" mo"', False),
        ("LTV / CAC", f"{e}!{{cl}}17", MULT_FMT, False),
        ("Magic number", f"{e}!{{cl}}18", MULT_FMT, False),
    ]

    row = HEADER_ROW + 1
    for label, formula, fmt, is_section in specs:
        if is_section:
            c = ws.cell(row, 1, value=label)
            c.font = FONT_BOLD
            row += 1
            continue
        _label(ws, row, label, indent=1)
        for i in range(len(YEARS)):
            col = DATA_COL0 + i
            cl = get_column_letter(col)
            ws.cell(row, col).value = "=" + formula.format(cl=cl)
            ws.cell(row, col).number_format = fmt
        row += 1

    _set_widths(ws, last_col)
    ws.sheet_view.showGridLines = False
    return name


def main() -> None:
    wb = load_workbook(WORKBOOK)
    src = wb[SOURCE_SHEET]
    data = _read_source(src)
    _normalize_source_data(src, data["cust_rows"])

    for stale in ("Sheet1", "KPI Summary", "ARR Bridge", "Cohort Retention",
                  "P&L", "Customer & Unit Economics"):
        if stale in wb.sheetnames:
            del wb[stale]

    bridge = build_arr_bridge(wb, data)
    cohort = build_cohort_retention(wb, data)
    pnl = build_pnl(wb, data)
    econ = build_customer_unit_econ(wb, data, bridge, pnl)
    kpi = build_kpi_summary(wb, bridge, pnl, econ)

    order = [kpi, bridge, cohort, pnl, econ, SOURCE_SHEET]
    by_title = {ws.title: ws for ws in wb.worksheets}
    wb._sheets = [by_title[t] for t in order if t in by_title] + [
        ws for ws in wb.worksheets if ws.title not in order
    ]

    wb.save(WORKBOOK)
    print(f"Wrote {WORKBOOK}")


if __name__ == "__main__":
    main()
